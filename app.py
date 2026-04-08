import streamlit as st
import pandas as pd
import numpy as np
import altair as alt
import io

# ─── Altair ダークテーマ ヘルパー ───
def _dk(chart):
    """Altairチャートにダークモード設定を適用"""
    return chart.configure(
        background="transparent",
        axis=alt.AxisConfig(
            labelColor="#8899bb", titleColor="#c8d8f0",
            gridColor="#1c2b44", domainColor="#3a4a6a",
            tickColor="#3a4a6a",
        ),
        legend=alt.LegendConfig(
            labelColor="#8899bb", titleColor="#c8d8f0",
            fillColor="#0d1526", strokeColor="#1c2b44",
        ),
        view=alt.ViewConfig(stroke="transparent"),
        title=alt.TitleConfig(color="#c8d8f0"),
    )

def build_excel_pl(df, df_yearly, rev_sources, fc_values_dict, vc_values_dict, industry_name, company_name="My Company"):
    """Generate a properly formatted P&L Excel file and return bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    NAVY   = "0F1E3C"
    AMBER  = "F59E0B"
    TEAL   = "14B8A6"
    WHITE  = "FFFFFF"
    LGRAY  = "F1F5F9"
    MGRAY  = "E2E8F0"
    DGRAY  = "64748B"
    GREEN  = "16A34A"
    RED    = "DC2626"

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ━━ SHEET 1: 月次P&L ━━
    ws = wb.active
    ws.title = "月次P&L"
    n_months = len(df)
    months = list(df["月番号"])

    ws.merge_cells(f"A1:{get_column_letter(n_months + 2)}1")
    c = ws["A1"]
    c.value = f"損益計算書 (P&L) — {company_name}  |  業種: {industry_name}"
    c.font = Font(bold=True, color=WHITE, size=13)
    c.fill = hdr_fill(NAVY)
    c.alignment = _align("left")

    ws.merge_cells(f"A2:{get_column_letter(n_months + 2)}2")
    c2 = ws["A2"]
    c2.value = f"シミュレーション期間: {n_months}ヶ月  |  BizMaker v7.0"
    c2.font = Font(color="94A3B8", size=9)
    c2.fill = hdr_fill("1A2F52")
    c2.alignment = _align("left")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16

    ws.cell(4, 1, "科目").font = Font(bold=True, color=WHITE, size=10)
    ws.cell(4, 1).fill = hdr_fill(NAVY)
    ws.cell(4, 2, "カテゴリ").font = Font(bold=True, color=WHITE, size=10)
    ws.cell(4, 2).fill = hdr_fill(NAVY)
    for i, m in enumerate(months):
        col = i + 3
        ws.cell(4, col, f"M{m}").font = Font(bold=True, color=WHITE, size=9)
        ws.cell(4, col).fill = hdr_fill(NAVY)
        ws.cell(4, col).alignment = _align("center")
    col_total = n_months + 3
    ws.cell(4, col_total, "合計").font = Font(bold=True, color=WHITE, size=10)
    ws.cell(4, col_total).fill = hdr_fill(AMBER)
    ws.cell(4, col_total).alignment = _align("center")
    ws.row_dimensions[4].height = 20

    _row = [5]
    def write_row(label, category, values, bold=False, bg=None, fg="1E293B", fmt='#,##0', indent=0):
        r = _row[0]
        ws.row_dimensions[r].height = 17
        label_text = ("  " * indent) + label
        c_l = ws.cell(r, 1, label_text)
        c_l.font = Font(bold=bold, color=fg, size=9)
        if bg: c_l.fill = hdr_fill(bg)
        c_l.alignment = _align("left")
        c_c = ws.cell(r, 2, category)
        c_c.font = Font(color=DGRAY, size=8)
        if bg: c_c.fill = hdr_fill(bg)
        total = 0
        for i, v in enumerate(values):
            col = i + 3
            cell = ws.cell(r, col, v)
            cell.number_format = fmt
            cell.font = Font(bold=bold, color=fg, size=9)
            cell.alignment = _align("right")
            if bg: cell.fill = hdr_fill(bg)
            total += v if v else 0
        tc = ws.cell(r, col_total, total)
        tc.number_format = fmt
        tc.font = Font(bold=bold, color=fg, size=9)
        tc.alignment = _align("right")
        if bg: tc.fill = hdr_fill(bg)
        _row[0] += 1

    def write_separator(bg=LGRAY):
        r = _row[0]
        ws.row_dimensions[r].height = 6
        for col in range(1, col_total + 1):
            ws.cell(r, col).fill = hdr_fill(bg)
        _row[0] += 1

    def write_section_header(label):
        r = _row[0]
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(r, 1, label)
        c.font = Font(bold=True, color=AMBER, size=9)
        c.fill = hdr_fill("1A2F52")
        for col in range(3, col_total + 1):
            ws.cell(r, col).fill = hdr_fill("1A2F52")
        _row[0] += 1

    write_section_header("▌ 売上高 (Revenue)")
    write_row("売上高合計", "Revenue", list(df["売上高"]), bold=True, bg="EFF6FF", fg="1E3A5F")
    write_separator()
    write_section_header("▌ 変動費 (Variable Costs)")
    write_row("変動費合計", "Variable Cost", list(df["変動費"]), bold=True, bg="FFF7ED", fg="7C2D12")
    write_row("  └ 広告宣伝費", "Advertising", list(df["広告宣伝費"]), fg=DGRAY, indent=1)
    write_separator()
    gp_vals = list(df["限界利益"])
    gp_rates = [gp / s * 100 if s > 0 else 0 for gp, s in zip(df["限界利益"], df["売上高"])]
    write_row("限界利益 (Gross Profit)", "Gross Profit", gp_vals, bold=True, bg="F0FDF4", fg="14532D")
    write_row("  └ 限界利益率 (%)", "Gross Margin%", gp_rates, bg="F0FDF4", fg=GREEN, fmt="0.0%")
    write_separator()
    write_section_header("▌ 販売費及び一般管理費 (SG&A)")
    write_row("固定費合計 (人件費込)", "Fixed + Labor", list(df["固定費合計"]), bold=True, fg="1E293B")
    for fc_name, fc_val in fc_values_dict.items():
        if isinstance(fc_val, (int, float)) and fc_val > 0:
            write_row(f"  └ {fc_name}", "Fixed Cost", [fc_val] * n_months, fg=DGRAY, indent=1)
    write_row("  └ 人件費 (採用計画分)", "Hire Cost", list(df["人件費（追加採用）"]), fg=DGRAY, indent=1)
    write_row("減価償却費", "D&A", list(df["減価償却費"]), fg="6D28D9")
    write_separator()
    ebitda_vals = [op + dep for op, dep in zip(df["営業利益"], df["減価償却費"])]
    write_row("EBITDA", "EBITDA", ebitda_vals, bold=True, bg="F5F3FF", fg="4C1D95")
    ebitda_rates = [e / s * 100 if s > 0 else 0 for e, s in zip(ebitda_vals, df["売上高"])]
    write_row("  └ EBITDA マージン (%)", "EBITDA Margin%", ebitda_rates, bg="F5F3FF", fg="7C3AED", fmt="0.0%")
    write_separator()
    write_row("営業利益 (EBIT)", "EBIT", list(df["営業利益"]), bold=True, bg="FFF1F2", fg="881337")
    op_rates = [op / s * 100 if s > 0 else 0 for op, s in zip(df["営業利益"], df["売上高"])]
    write_row("  └ 営業利益率 (%)", "EBIT Margin%", op_rates, bg="FFF1F2", fg=RED, fmt="0.0%")
    write_row("法人税等", "Tax", list(df["税額"]), fg=DGRAY)
    write_row("税シールド (減価償却)", "Tax Shield", list(df["税シールド"]), fg=TEAL)
    write_separator()
    write_row("当期純利益 (Net Income)", "Net Income", list(df["税引後利益"]), bold=True, bg="F0FDF4", fg="14532D")
    ni_rates = [ni / s * 100 if s > 0 else 0 for ni, s in zip(df["税引後利益"], df["売上高"])]
    write_row("  └ 純利益率 (%)", "Net Margin%", ni_rates, bg="F0FDF4", fg=GREEN, fmt="0.0%")
    write_separator()
    write_section_header("▌ 累積・キャッシュフロー")
    write_row("累積利益", "Cum Profit", list(df["累積利益"]), bold=True)
    write_row("累積税引後利益", "Cum Net Income", list(df["累積税引後利益"]), bold=True)
    write_row("キャッシュ残高", "Cash Balance", list(df["キャッシュ残高"]), bold=True, fg="0369A1")
    write_separator()
    write_section_header("▌ 営業指標 (KPIs)")
    write_row("新規獲得顧客数", "New Users", list(df["新規獲得"]), fmt="#,##0")
    write_row("アクティブ顧客数", "Active Users", list(df["アクティブ顧客数"]), fmt="#,##0")
    write_row("解約数", "Churn", list(df["解約数"]), fmt="#,##0")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    for i in range(3, col_total + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.freeze_panes = "C5"

    # ━━ SHEET 2: 年次サマリー ━━
    ws2 = wb.create_sheet("年次サマリー")
    n_years = len(df_yearly)
    yr_col_total = n_years + 3
    ws2.merge_cells(f"A1:{get_column_letter(yr_col_total)}1")
    ws2["A1"].value = f"年次損益サマリー — {company_name}"
    ws2["A1"].font = Font(bold=True, color=WHITE, size=12)
    ws2["A1"].fill = hdr_fill(NAVY)
    ws2["A1"].alignment = _align("left")
    ws2.row_dimensions[1].height = 22
    ws2.cell(2, 1, "科目").font = Font(bold=True, color=WHITE, size=10)
    ws2.cell(2, 1).fill = hdr_fill(NAVY)
    ws2.cell(2, 2, "カテゴリ").font = Font(bold=True, color=WHITE, size=10)
    ws2.cell(2, 2).fill = hdr_fill(NAVY)
    for i, yr_val in enumerate(df_yearly["年"]):
        col = i + 3
        ws2.cell(2, col, f"Year {int(yr_val)}").font = Font(bold=True, color=WHITE, size=10)
        ws2.cell(2, col).fill = hdr_fill(NAVY)
        ws2.cell(2, col).alignment = _align("center")
    ws2.cell(2, yr_col_total, "累計").font = Font(bold=True, color=WHITE, size=10)
    ws2.cell(2, yr_col_total).fill = hdr_fill(AMBER)
    ws2.cell(2, yr_col_total).alignment = _align("center")
    ws2.row_dimensions[2].height = 20

    def write_yr_row(ws_t, yr_r, label, category, values, bold=False, bg=None, fg="1E293B", fmt='#,##0'):
        ws_t.row_dimensions[yr_r].height = 17
        c = ws_t.cell(yr_r, 1, label)
        c.font = Font(bold=bold, color=fg, size=9)
        if bg: c.fill = hdr_fill(bg)
        c.alignment = _align("left")
        c2x = ws_t.cell(yr_r, 2, category)
        c2x.font = Font(color=DGRAY, size=8)
        if bg: c2x.fill = hdr_fill(bg)
        total = 0
        for i, v in enumerate(values):
            col = i + 3
            cell = ws_t.cell(yr_r, col, v)
            cell.number_format = fmt
            cell.font = Font(bold=bold, color=fg, size=9)
            cell.alignment = _align("right")
            if bg: cell.fill = hdr_fill(bg)
            total += v if v else 0
        tc = ws_t.cell(yr_r, yr_col_total, total)
        tc.number_format = fmt
        tc.font = Font(bold=bold, color=fg, size=9)
        tc.alignment = _align("right")
        if bg: tc.fill = hdr_fill(bg)

    yr = 3
    write_yr_row(ws2, yr, "売上高", "Revenue", list(df_yearly["売上高"]), bold=True, bg="EFF6FF", fg="1E3A5F"); yr += 1
    write_yr_row(ws2, yr, "変動費", "Variable Cost", list(df_yearly["変動費"]), fg="7C2D12"); yr += 1
    write_yr_row(ws2, yr, "限界利益", "Gross Profit", list(df_yearly["限界利益"]), bold=True, bg="F0FDF4", fg="14532D"); yr += 1
    gm_rates_y = [gp / s * 100 if s > 0 else 0 for gp, s in zip(df_yearly["限界利益"], df_yearly["売上高"])]
    write_yr_row(ws2, yr, "  └ 限界利益率 (%)", "Gross Margin%", gm_rates_y, fg=GREEN, fmt="0.0%"); yr += 1
    write_yr_row(ws2, yr, "固定費合計", "Fixed+Labor", list(df_yearly["固定費合計"]), fg=DGRAY); yr += 1
    write_yr_row(ws2, yr, "広告宣伝費", "Advertising", list(df_yearly["広告宣伝費"]), fg=DGRAY); yr += 1
    write_yr_row(ws2, yr, "減価償却費", "D&A", list(df_yearly["減価償却費"])); yr += 1
    ebitda_yr = [op + dep for op, dep in zip(df_yearly["営業利益"], df_yearly["減価償却費"])]
    write_yr_row(ws2, yr, "EBITDA", "EBITDA", ebitda_yr, bold=True, bg="F5F3FF", fg="4C1D95"); yr += 1
    write_yr_row(ws2, yr, "営業利益", "EBIT", list(df_yearly["営業利益"]), bold=True, bg="FFF1F2", fg="881337"); yr += 1
    op_rates_yr = [op / s * 100 if s > 0 else 0 for op, s in zip(df_yearly["営業利益"], df_yearly["売上高"])]
    write_yr_row(ws2, yr, "  └ 営業利益率 (%)", "EBIT Margin%", op_rates_yr, fg=RED, fmt="0.0%"); yr += 1
    write_yr_row(ws2, yr, "税額", "Tax", list(df_yearly["税額"]), fg=DGRAY); yr += 1
    write_yr_row(ws2, yr, "当期純利益", "Net Income", list(df_yearly["税引後利益"]), bold=True, bg="F0FDF4", fg="14532D"); yr += 1
    ni_rates_yr = [ni / s * 100 if s > 0 else 0 for ni, s in zip(df_yearly["税引後利益"], df_yearly["売上高"])]
    write_yr_row(ws2, yr, "  └ 純利益率 (%)", "Net Margin%", ni_rates_yr, fg=GREEN, fmt="0.0%"); yr += 1
    write_yr_row(ws2, yr, "キャッシュ残高 (期末)", "Cash (EOP)", list(df_yearly["キャッシュ残高"]), bold=True, fg="0369A1"); yr += 1
    write_yr_row(ws2, yr, "アクティブ顧客数 (期末)", "MAU (EOP)", list(df_yearly["アクティブ顧客数"].astype(int)), fmt="#,##0"); yr += 1
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    for i in range(3, yr_col_total + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.freeze_panes = "C3"

    # ━━ SHEET 3: KPIサマリー ━━
    ws3 = wb.create_sheet("KPIサマリー")
    ws3["A1"] = "KPIサマリー"
    ws3["A1"].font = Font(bold=True, color=WHITE, size=12)
    ws3["A1"].fill = hdr_fill(NAVY)
    ws3.row_dimensions[1].height = 20
    kpi_rows_data = [
        ("業種", industry_name),
        ("シミュレーション期間", f"{len(df)}ヶ月"),
        ("最終月売上高", df["売上高"].iloc[-1]),
        ("最終月営業利益", df["営業利益"].iloc[-1]),
        ("最終月純利益", df["税引後利益"].iloc[-1]),
        ("最終月キャッシュ残高", df["キャッシュ残高"].iloc[-1]),
        ("最終月アクティブ顧客数", df["アクティブ顧客数"].iloc[-1]),
        ("累積利益合計", df["累積利益"].iloc[-1]),
        ("累積純利益合計", df["累積税引後利益"].iloc[-1]),
        ("月次黒字化月", next((m for m in df["月番号"] if df.loc[df["月番号"]==m, "営業利益"].values[0] > 0), "期間外")),
    ]
    for i, (k, v) in enumerate(kpi_rows_data, 2):
        ws3.cell(i, 1, k).font = Font(color=DGRAY, size=9)
        cell = ws3.cell(i, 2, v)
        if isinstance(v, float):
            cell.number_format = "#,##0"
        cell.font = Font(bold=True, size=10)
        ws3.row_dimensions[i].height = 16
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Biz Maker — ビジネス共創プラットフォーム",
    layout="wide",
    page_icon="🚀",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────
# GLOBAL CSS — Dark Mode (Plan A: Deep Navy × Amber)
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
html, body, [class*="css"] { font-family:'Inter',sans-serif !important; }
#MainMenu, header, footer { visibility:hidden; }

/* ── Global Dark Background ── */
.main .block-container { padding:1.2rem 2rem 2rem; background:#0b1220; }
body, .stApp { background:#0b1220 !important; }
section[data-testid="stSidebar"] { background:#0d1526 !important; }

/* ── ナビバー ── */
.top-nav { display:flex; align-items:center; justify-content:space-between; padding:0.6rem 0; margin-bottom:1rem; border-bottom:1px solid #1c2b44; }
.top-nav .logo { font-size:2.0rem; font-weight:800; color:#fff; letter-spacing:-0.5px; }
.top-nav .logo span { color:#f5a623; }
.top-nav .tagline { font-size:0.85rem; color:#8899bb; margin-top:2px; }
.nav-badge { background:#1c2b44; color:#f5a623; border-radius:20px; padding:2px 10px; font-size:0.7rem; font-weight:700; border:1px solid rgba(245,166,35,.35); }

/* ── KPIカード ── */
.kpi-grid { display:flex; gap:12px; flex-wrap:wrap; margin:1rem 0; }
.kpi-card { flex:1; min-width:140px; background:#0d1526; border:1px solid #1c2b44; border-radius:12px; padding:14px 16px; }
.kpi-card .label { font-size:0.72rem; color:#8899bb; font-weight:500; text-transform:uppercase; letter-spacing:0.5px; }
.kpi-card .value { font-size:1.35rem; font-weight:700; color:#fff; margin:4px 0 2px; line-height:1.2; }
.kpi-card .delta { font-size:0.75rem; }
.kpi-card .delta.up   { color:#34d297; }
.kpi-card .delta.down { color:#f87171; }
.kpi-card .delta.neutral { color:#8899bb; }
.kpi-card.accent  { border-left:3px solid #f5a623; }
.kpi-card.success { border-left:3px solid #34d297; }
.kpi-card.danger  { border-left:3px solid #f87171; }
.kpi-card.warn    { border-left:3px solid #f5a623; }

/* ── セクションタイトル ── */
.section-title { font-size:0.8rem; font-weight:700; color:#f5a623; text-transform:uppercase; letter-spacing:1px; margin:1.4rem 0 0.6rem; padding-bottom:6px; border-bottom:1px solid #1c2b44; }

/* ── 改善提案カード ── */
.advice-card { background:#0d1526; border:1px solid #1c2b44; border-radius:10px; padding:14px 16px; margin:6px 0; }
.advice-card .advice-title { font-size:0.8rem; font-weight:600; color:#b0c0d8; margin-bottom:4px; }
.advice-card .advice-value { font-size:1.05rem; font-weight:700; color:#fff; }
.advice-card .advice-desc { font-size:0.75rem; color:#8899bb; margin-top:3px; }

/* ── ステップバー ── */
.step-bar { display:flex; gap:0; margin:0.8rem 0 1.2rem; }
.step-item { flex:1; text-align:center; padding:8px 4px; font-size:0.72rem; font-weight:500; border-bottom:3px solid #1c2b44; color:#8899bb; }
.step-item.active { border-bottom:3px solid #f5a623; color:#f5a623; font-weight:700; }
.step-item.done   { border-bottom:3px solid #34d297; color:#34d297; }

/* ── ファネルビジュアル ── */
.funnel-total { background:#0d1526; border:1px solid #f5a623; border-radius:12px; padding:16px 20px; margin:10px 0; }
.funnel-total-title { font-size:0.75rem; font-weight:700; color:#f5a623; margin-bottom:12px; letter-spacing:1px; }
.funnel-row { display:flex; align-items:center; gap:0; }
.funnel-step { flex:1; text-align:center; }
.funnel-num  { font-size:1.3rem; font-weight:800; font-family:monospace; }
.funnel-lbl  { font-size:0.65rem; color:#8899bb; margin-top:2px; }
.funnel-rate { font-size:0.68rem; font-weight:600; margin-top:2px; }
.funnel-arr  { color:#6a8aaa; font-size:1.1rem; padding:0 6px; flex-shrink:0; }

/* ── チャネルファネル（各チャネル） ── */
.ch-funnel { display:flex; align-items:center; background:#111e32; border:1px solid #1c2b44; border-radius:8px; padding:10px 14px; margin:6px 0; font-family:monospace; gap:0; }
.ch-funnel .cf-step { flex:1; text-align:center; }
.ch-funnel .cf-num { font-size:0.95rem; font-weight:700; }
.ch-funnel .cf-lbl { font-size:0.6rem; color:#8899bb; margin-top:1px; }
.ch-funnel .cf-sub { font-size:0.6rem; color:#8899bb; margin-top:1px; }
.ch-funnel .cf-arr { color:#6a8aaa; padding:0 5px; flex-shrink:0; }

/* ── AIチャット ── */
.ai-bubble { background:linear-gradient(135deg,#1c2b44,#162236); color:#c8d8f0; border-radius:0 14px 14px 14px; padding:14px 18px; margin:8px 0 8px 8px; font-size:0.88rem; line-height:1.65; border:1px solid rgba(245,166,35,.15); max-width:85%; }
.ai-label { font-size:0.7rem; font-weight:700; color:#f5a623; margin-bottom:6px; letter-spacing:0.5px; }
.user-bubble { background:#131e32; color:#c8d8f0; border-radius:14px 0 14px 14px; padding:12px 16px; margin:8px 8px 8px auto; font-size:0.87rem; max-width:75%; text-align:right; float:right; clear:both; border:1px solid #1c2b44; }
.chat-wrap { overflow:hidden; }

/* ── コンサルカード ── */
.consultant-card { background:#0d1526; border:1px solid #1c2b44; border-radius:14px; padding:20px; margin:10px 0; transition:all 0.2s; }
.consultant-card:hover { border-color:rgba(245,166,35,.5); box-shadow:0 0 20px rgba(245,166,35,.08); }
.cons-name { font-size:1.0rem; font-weight:700; color:#fff; }
.cons-field { font-size:0.75rem; color:#f5a623; font-weight:600; background:rgba(245,166,35,.1); padding:2px 8px; border-radius:20px; border:1px solid rgba(245,166,35,.2); }
.cons-desc { font-size:0.82rem; color:#a0b0c8; margin:8px 0; line-height:1.5; }
.cons-meta { font-size:0.78rem; color:#8899bb; }
.cons-badge { background:rgba(52,210,151,.1); color:#34d297; border-radius:20px; padding:2px 10px; font-size:0.72rem; font-weight:600; border:1px solid rgba(52,210,151,.2); display:inline-block; margin:4px 0; }

/* ── SNS投稿カード ── */
.post-card { background:#0d1526; border:1px solid #1c2b44; border-radius:12px; padding:16px; margin:10px 0; }
.post-header { display:flex; align-items:center; gap:10px; margin-bottom:10px; }
.post-avatar { width:36px; height:36px; border-radius:50%; background:linear-gradient(135deg,#1c2b44,#f5a623); display:flex; align-items:center; justify-content:center; color:#fff; font-size:0.85rem; font-weight:700; }
.post-meta .name { font-weight:600; font-size:0.88rem; color:#fff; }
.post-meta .sub { font-size:0.72rem; color:#8899bb; }
.post-content { font-size:0.85rem; color:#b0c0d8; line-height:1.65; }
.post-actions { display:flex; gap:20px; margin-top:12px; padding-top:10px; border-top:1px solid #1c2b44; }
.post-action { font-size:0.78rem; color:#8899bb; cursor:pointer; font-weight:500; }
.tag { background:rgba(245,166,35,.1); color:#f5a623; border:1px solid rgba(245,166,35,.2); border-radius:4px; padding:1px 6px; font-size:0.7rem; font-weight:600; display:inline-block; margin:2px; }

/* ── Coming Soonバッジ ── */
.cs-banner { background:rgba(245,166,35,.06); border:1px solid rgba(245,166,35,.2); border-radius:10px; padding:10px 16px; margin-bottom:16px; font-size:0.82rem; color:#b0c0d8; font-weight:500; }

/* ── Metric override ── */
div[data-testid="stMetric"] { background:#0d1526 !important; border-radius:10px; padding:12px 14px; border:1px solid #1c2b44 !important; }
div[data-testid="stMetric"] label { color:#a0b0c8 !important; }
div[data-testid="stMetric"] div[data-testid="stMetricValue"] { color:#fff !important; }

/* ── Button ── */
div[data-testid="stButton"] > button { border-radius:8px !important; font-weight:600 !important; font-size:0.85rem !important; padding:6px 14px !important; background:#1c2b44 !important; color:#c8d8f0 !important; border:1px solid #2a3d58 !important; transition:all 0.2s !important; }
div[data-testid="stButton"] > button:hover { background:#253d5c !important; border-color:#f5a623 !important; color:#f5a623 !important; }

/* ── Tab Orange Pill Style ── */
div[data-testid="stTabs"] [data-baseweb="tab-list"] { gap:14px; border-bottom:none !important; background:transparent; padding:10px 0; }
div[data-testid="stTabs"] [data-baseweb="tab"] {
    background:linear-gradient(180deg,#F97316 0%,#EA580C 100%); color:#fff !important;
    border-radius:50px; padding:14px 32px; font-weight:700; font-size:0.95rem;
    border:none !important; letter-spacing:0.5px;
    box-shadow:0 4px 14px rgba(234,88,12,0.35), inset 0 1px 0 rgba(255,255,255,0.25);
    transition:all 0.25s ease; cursor:pointer; position:relative; overflow:hidden;
}
div[data-testid="stTabs"] [data-baseweb="tab"]::before { content:""; position:absolute; top:0; left:0; right:0; bottom:0; background:linear-gradient(180deg,rgba(255,255,255,0.12) 0%,transparent 60%); border-radius:50px; pointer-events:none; }
div[data-testid="stTabs"] [data-baseweb="tab"]:hover { background:linear-gradient(180deg,#FB923C 0%,#F97316 100%); box-shadow:0 6px 22px rgba(249,115,22,0.45); transform:translateY(-2px); }
div[data-testid="stTabs"] [data-baseweb="tab"][aria-selected="true"] { background:linear-gradient(180deg,#EA580C 0%,#C2410C 100%) !important; box-shadow:0 2px 8px rgba(194,65,12,0.4), inset 0 2px 4px rgba(0,0,0,0.15) !important; transform:translateY(1px); }
div[data-testid="stTabs"] [data-baseweb="tab-highlight"], div[data-testid="stTabs"] [data-baseweb="tab-border"] { display:none !important; }
div[data-testid="stTabs"] [data-baseweb="tab"] > div { color:#fff !important; }

/* ── Input dark override ── */
div[data-baseweb="input"] > div { background:#111e32 !important; border-color:#1c2b44 !important; }
div[data-baseweb="input"] input { color:#c8d8f0 !important; }
div[data-baseweb="select"] > div { background:#111e32 !important; border-color:#1c2b44 !important; }
div[data-baseweb="select"] span { color:#c8d8f0 !important; }
textarea { background:#111e32 !important; color:#c8d8f0 !important; border-color:#1c2b44 !important; }
div[data-testid="stNumberInput"] input { color:#c8d8f0 !important; }
div[data-testid="stExpander"] { background:#0d1526 !important; border:1px solid #1c2b44 !important; border-radius:10px !important; }
div[data-testid="stExpander"] > details > summary { color:#c8d8f0 !important; }
div[data-testid="stExpander"] > details > summary:hover { color:#f5a623 !important; }
.stCheckbox label span { color:#c8d8f0 !important; }
.stRadio label { color:#c8d8f0 !important; }
div[data-testid="stDataFrame"] { background:#0d1526 !important; }
div[role="listbox"] { background:#111e32 !important; border:1px solid #1c2b44 !important; }

/* ── Text dark ── */
h1, h2, h3, h4, h5, h6 { color:#fff !important; }
.stMarkdown p { color:#8899bb; }
.stCaption { color:#8899bb !important; }
label[data-testid="stWidgetLabel"] { color:#8899bb !important; }

/* ── 資金アラート ── */
.funding-alert { background:rgba(248,113,113,.07); border:1px solid rgba(248,113,113,.3); border-radius:12px; padding:16px 20px; margin:12px 0; }
.funding-alert .fa-title { font-size:0.9rem; font-weight:700; color:#f87171; margin-bottom:4px; }
.funding-alert .fa-body { font-size:0.82rem; color:#8899bb; line-height:1.6; }

/* ── Growth Navigator ── */
.gn-goal-card {
    background:linear-gradient(135deg,#0d1526,#111e32);
    border:1px solid #1c2b44; border-radius:12px; padding:18px 20px; margin:8px 0;
    transition: border-color 0.2s;
}
.gn-goal-card:hover { border-color:#f5a623; }
.gn-goal-card .gn-title { font-size:0.95rem; font-weight:700; color:#f5a623; margin-bottom:6px; }
.gn-goal-card .gn-value { font-size:1.4rem; font-weight:800; color:#fff; }
.gn-goal-card .gn-sub { font-size:0.78rem; color:#a0b0c8; margin-top:4px; }
.gn-lever-row {
    background:#0d1526; border:1px solid #1c2b44; border-radius:10px;
    padding:14px 18px; margin:6px 0; display:flex; align-items:center; gap:14px;
}
.gn-lever-row .lever-rank { font-size:1.2rem; font-weight:800; color:#f5a623; min-width:28px; }
.gn-lever-row .lever-name { font-size:0.88rem; font-weight:600; color:#c8d8f0; }
.gn-lever-row .lever-impact { font-size:0.78rem; color:#4ade80; }
.gn-lever-row .lever-current { font-size:0.78rem; color:#a0b0c8; }
.gn-benchmark-bar { height:8px; border-radius:4px; background:#1c2b44; position:relative; margin:4px 0; }
.gn-benchmark-fill { height:8px; border-radius:4px; position:absolute; left:0; top:0; }
.gn-section-header { font-size:1.05rem; font-weight:700; color:#fff; margin:20px 0 10px; padding-bottom:6px; border-bottom:2px solid #f5a62344; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 業種テンプレート（A4 拡張：選択式コスト項目）
# ─────────────────────────────────────────────
# 変動費マスター：全業種共通のコスト項目プール
VARIABLE_COST_ITEMS = {
    "仕入原価":       {"key": "vc_cogs",     "unit": "円/件", "desc": "商品の仕入れ・製造原価"},
    "配送料":         {"key": "vc_shipping",  "unit": "円/件", "desc": "配送・物流費用"},
    "サーバー原価":   {"key": "vc_server",    "unit": "円/件", "desc": "SaaS・クラウドの従量課金"},
    "決済手数料":     {"key": "vc_payment",   "unit": "%",     "desc": "クレジットカード等の決済手数料"},
    "モール手数料":   {"key": "vc_platform",  "unit": "%",     "desc": "ECモール等のプラットフォーム手数料"},
    "外注加工費":     {"key": "vc_outsource", "unit": "円/件", "desc": "外部への加工・制作委託"},
    "梱包資材費":     {"key": "vc_packaging", "unit": "円/件", "desc": "梱包材・パッケージ費用"},
    "ロイヤリティ":   {"key": "vc_royalty",   "unit": "%",     "desc": "ライセンス・ロイヤリティ費用"},
    "販売手数料":     {"key": "vc_sales_fee", "unit": "%",     "desc": "販売代理店・アフィリエイト手数料"},
    "返品コスト":     {"key": "vc_returns",   "unit": "円/件", "desc": "返品・交換に伴う費用"},
}

FIXED_COST_ITEMS = {
    "給与合計":       {"key": "fc_salary",      "desc": "正社員・パート給与の合計"},
    "社会保険料":     {"key": "fc_insurance",    "desc": "健康保険・厚生年金等"},
    "業務委託費":     {"key": "fc_outsourcing",  "desc": "フリーランス・外注への固定払い"},
    "家賃":           {"key": "fc_rent",         "desc": "オフィス・店舗の賃料"},
    "システム利用料": {"key": "fc_system",       "desc": "SaaS月額・ツール利用料"},
    "その他固定費":   {"key": "fc_misc",         "desc": "雑費・交際費・通信費等"},
    "リース料":       {"key": "fc_lease",        "desc": "設備・車両リース料"},
    "広告宣伝費（固定）": {"key": "fc_ad_fixed", "desc": "ブランディング・PR等の固定広告費"},
    "研究開発費":     {"key": "fc_rd",           "desc": "R&D・新規開発の固定投資"},
    "保険料":         {"key": "fc_business_ins", "desc": "事業保険・賠償保険等"},
    "水道光熱費":     {"key": "fc_utilities",    "desc": "電気・ガス・水道料金"},
    "通信費":         {"key": "fc_telecom",      "desc": "電話・インターネット回線費"},
    "交通費":         {"key": "fc_transport",    "desc": "出張・通勤交通費"},
    "顧問料":         {"key": "fc_advisory",     "desc": "税理士・弁護士等の顧問契約"},
    "採用費":         {"key": "fc_recruiting",   "desc": "求人広告・人材紹介手数料"},
}

INDUSTRY_TEMPLATES = {
    "カスタム": {
        "unit_price": 5000, "ad_budget": 1_000_000, "cpa": 2000,
        "organic_start": 50, "organic_growth": 5.0, "churn_rate": 5.0,
        "vc_items": {"仕入原価": 1000, "配送料": 600, "サーバー原価": 50, "決済手数料": 3.6, "モール手数料": 0.0},
        "fc_items": {"給与合計": 2_500_000, "社会保険料": 400_000, "業務委託費": 300_000, "家賃": 150_000, "システム利用料": 50_000, "その他固定費": 100_000},
        "seasonal": [1.0]*12,
    },
    "SaaS / サブスク": {
        "unit_price": 9800, "ad_budget": 800_000, "cpa": 8000,
        "organic_start": 30, "organic_growth": 8.0, "churn_rate": 3.0,
        "vc_items": {"サーバー原価": 200, "決済手数料": 3.6},
        "fc_items": {"給与合計": 3_000_000, "社会保険料": 450_000, "業務委託費": 500_000, "家賃": 100_000, "システム利用料": 150_000, "その他固定費": 100_000},
        "seasonal": [1.0,1.0,1.05,1.0,1.0,1.0,0.95,0.95,1.05,1.05,1.0,0.95],
    },
    "EC / 通販": {
        "unit_price": 4500, "ad_budget": 1_500_000, "cpa": 2500,
        "organic_start": 100, "organic_growth": 3.0, "churn_rate": 15.0,
        "vc_items": {"仕入原価": 1800, "配送料": 800, "サーバー原価": 30, "決済手数料": 3.6, "モール手数料": 8.0, "梱包資材費": 150},
        "fc_items": {"給与合計": 2_000_000, "社会保険料": 300_000, "業務委託費": 200_000, "家賃": 200_000, "システム利用料": 80_000, "その他固定費": 120_000},
        "seasonal": [0.8,0.8,1.0,0.9,0.9,1.0,1.1,0.9,0.9,1.0,1.2,1.5],
    },
    "飲食店": {
        "unit_price": 1200, "ad_budget": 300_000, "cpa": 500,
        "organic_start": 200, "organic_growth": 2.0, "churn_rate": 25.0,
        "vc_items": {"仕入原価": 400, "決済手数料": 3.6, "梱包資材費": 30},
        "fc_items": {"給与合計": 1_800_000, "社会保険料": 270_000, "業務委託費": 50_000, "家賃": 300_000, "システム利用料": 30_000, "その他固定費": 150_000},
        "seasonal": [0.8,0.85,1.0,1.0,1.0,0.9,0.95,0.85,0.95,1.0,1.1,1.5],
    },
    "コンサルティング": {
        "unit_price": 300_000, "ad_budget": 500_000, "cpa": 50_000,
        "organic_start": 5, "organic_growth": 5.0, "churn_rate": 8.0,
        "vc_items": {"決済手数料": 3.6},
        "fc_items": {"給与合計": 3_500_000, "社会保険料": 525_000, "業務委託費": 200_000, "家賃": 200_000, "システム利用料": 50_000, "その他固定費": 100_000},
        "seasonal": [0.7,0.8,1.2,1.1,1.0,1.0,0.9,0.7,1.0,1.1,1.1,1.2],
    },
    "ハードウェア": {
        "unit_price": 25_000, "ad_budget": 2_000_000, "cpa": 5000,
        "organic_start": 30, "organic_growth": 3.0, "churn_rate": 0.0,
        "vc_items": {"仕入原価": 10_000, "配送料": 1500, "決済手数料": 3.6, "梱包資材費": 500},
        "fc_items": {"給与合計": 4_000_000, "社会保険料": 600_000, "業務委託費": 1_000_000, "家賃": 500_000, "システム利用料": 100_000, "その他固定費": 300_000},
        "seasonal": [0.9,0.8,1.0,1.0,1.0,1.0,1.0,0.9,1.0,1.0,1.1,1.3],
    },
    "買い切り＋サブスク": {
        "unit_price": 35_000, "ad_budget": 1_200_000, "cpa": 4000,
        "organic_start": 20, "organic_growth": 4.0, "churn_rate": 5.0,
        "vc_items": {"仕入原価": 12_000, "配送料": 1_000, "サーバー原価": 100, "決済手数料": 3.6},
        "fc_items": {"給与合計": 3_000_000, "社会保険料": 450_000, "業務委託費": 500_000, "家賃": 200_000, "システム利用料": 100_000, "その他固定費": 150_000},
        "seasonal": [0.9,0.85,1.0,1.0,1.0,1.0,1.0,0.9,1.0,1.05,1.1,1.2],
    },
}

# ─────────────────────────────────────────────
# 業種ベンチマーク (Growth Navigator用)
# ─────────────────────────────────────────────
INDUSTRY_BENCHMARKS = {
    "SaaS / サブスク":      {"CTR": 2.5, "CVR": 5.0, "churn": 3.0, "LTV_CAC": 3.0, "ARPU": 9800,  "gross_margin": 80},
    "EC / 通販":            {"CTR": 1.8, "CVR": 3.0, "churn": 15.0,"LTV_CAC": 2.5, "ARPU": 4500,  "gross_margin": 40},
    "飲食店":               {"CTR": 3.5, "CVR": 8.0, "churn": 25.0,"LTV_CAC": 2.0, "ARPU": 1200,  "gross_margin": 65},
    "コンサルティング":     {"CTR": 1.2, "CVR": 2.0, "churn": 8.0, "LTV_CAC": 5.0, "ARPU": 300000,"gross_margin": 85},
    "ハードウェア":         {"CTR": 1.5, "CVR": 2.5, "churn": 0.0, "LTV_CAC": 2.0, "ARPU": 25000, "gross_margin": 35},
    "買い切り＋サブスク":   {"CTR": 2.0, "CVR": 3.5, "churn": 5.0, "LTV_CAC": 2.8, "ARPU": 35000, "gross_margin": 50},
    "メディア / 広告":      {"CTR": 2.0, "CVR": 4.0, "churn": 12.0,"LTV_CAC": 2.5, "ARPU": 500,   "gross_margin": 70},
    "教育 / EdTech":        {"CTR": 2.2, "CVR": 4.5, "churn": 8.0, "LTV_CAC": 3.0, "ARPU": 6000,  "gross_margin": 75},
    "ヘルスケア":           {"CTR": 1.5, "CVR": 3.0, "churn": 6.0, "LTV_CAC": 3.5, "ARPU": 12000, "gross_margin": 60},
    "フィンテック":         {"CTR": 1.0, "CVR": 2.0, "churn": 4.0, "LTV_CAC": 4.0, "ARPU": 15000, "gross_margin": 70},
    "カスタム":             {"CTR": 2.0, "CVR": 4.0, "churn": 5.0, "LTV_CAC": 3.0, "ARPU": 5000,  "gross_margin": 50},
}

MONTH_LABELS = ["1月","2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月"]

# ─── 減価償却の種別定義 (A3) ───
DEPRECIATION_CATEGORIES = {
    "設備・機械": {"useful_life": 7, "examples": "製造機械、加工設備、厨房設備"},
    "IT資産": {"useful_life": 4, "examples": "PC、サーバー、ネットワーク機器"},
    "車両": {"useful_life": 6, "examples": "営業車、配送トラック"},
    "不動産（建物）": {"useful_life": 22, "examples": "店舗内装、オフィス内装"},
    "ソフトウェア": {"useful_life": 5, "examples": "自社開発ソフト、ライセンス"},
    "その他": {"useful_life": 5, "examples": "工具、備品、什器"},
}

# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
defaults = {
    "step": 1,
    "industry": "SaaS / サブスク",
    "revenue_sources": [{"name": "メイン商品", "unit_price": 9800, "weight": 100}],
    "hire_plan": [],
    "depreciation_assets": [],
    "acq_mode": "funnel",
    "channels": [
        {"name": "リスティング広告", "budget": 500_000, "cpm": 600, "ctr": 2.5, "cvr": 2.5},
    ],
}
for key, val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = val

# ─────────────────────────────────────────────
# TOP NAV
# ─────────────────────────────────────────────
st.markdown("""
<div class="top-nav">
  <div>
    <div class="logo">Biz<span>Maker</span></div>
    <div class="tagline">ビジネス共創プラットフォーム</div>
  </div>
  <div><span class="nav-badge">v7.0 — Growth Navigator</span></div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# MAIN TABS
# ─────────────────────────────────────────────
tab_sim, tab_ai, tab_cons, tab_sns = st.tabs([
    " シミュレーター ", " AI アドバイザー ",
    " 専門家に相談 ", " コミュニティ ",
])


# ═══════════════════════════════════════════════
# TAB 1 — シミュレーター
# ═══════════════════════════════════════════════
with tab_sim:
    # ── 業種テンプレート & シナリオ ──
    col_ind, col_sc = st.columns([3, 1])
    with col_ind:
        industry = st.selectbox(
            "業種テンプレート",
            list(INDUSTRY_TEMPLATES.keys()),
            index=list(INDUSTRY_TEMPLATES.keys()).index(st.session_state.industry),
        )
        st.session_state.industry = industry
        tmpl = INDUSTRY_TEMPLATES[industry]
    with col_sc:
        scenario = st.radio("シナリオ", ["中庸", "楽観 +20%", "悲観 -20%"], horizontal=False)

    # ── ステップバー (7ステップに拡張) ──
    step_labels = [
        "Step 1 基本情報",
        "Step 2 売上設計",
        "Step 3 コスト設計",
        "Step 4 人員計画",
        "Step 5 設備投資",
        "Step 6 資金繰り",
        "Step 7 税効果",
    ]
    step_html = '<div class="step-bar">'
    for i, sl in enumerate(step_labels, 1):
        cls = "active" if i == st.session_state.step else ("done" if i < st.session_state.step else "")
        step_html += f'<div class="step-item {cls}">{sl}</div>'
    step_html += "</div>"
    st.markdown(step_html, unsafe_allow_html=True)

    # ═══════════════════════════════════════
    # STEP 1 — 基本情報 (A2 + A5)
    # ═══════════════════════════════════════
    with st.expander("Step 1 — 基本情報", expanded=(st.session_state.step == 1)):
        c1, c2, c3 = st.columns(3)
        with c1:
            sim_months = st.selectbox("シミュレーション期間", [12, 24, 36, 60, 84, 120], index=2)
        with c2:
            target_pct = st.slider("目標営業利益率 (%)", 1, 50, 20)
            target_rate = target_pct / 100
        with c3:
            initial_inv = st.number_input("初期投資額 (円)", value=5_000_000, step=500_000)

        # A5: 入力粒度の選択
        input_mode = st.radio(
            "入力モード",
            ["月次入力（詳細）", "年次入力（概算）"],
            horizontal=True,
            help="年次入力を選ぶと、売上・コストを年額で入力し自動で月割り計算します",
        )
        is_annual_input = input_mode == "年次入力（概算）"
        annual_divisor = 12 if is_annual_input else 1
        annual_label = " (年額)" if is_annual_input else ""

        _, nb = st.columns([8, 1])
        with nb:
            if st.button("次へ", key="n1"):
                st.session_state.step = 2; st.rerun()

    # ═══════════════════════════════════════
    # STEP 2 — 売上設計 (B2: 複数収益源 + IMPROVEMENT 5: 年度別成長率)
    # ═══════════════════════════════════════
    with st.expander("Step 2 — 売上設計（複数収益源対応）", expanded=(st.session_state.step == 2)):
        st.caption("収益源を追加して、複数のプロダクト・サービスの売上を個別にシミュレーションできます。")

        # 収益源の数を管理
        if "n_revenue" not in st.session_state:
            st.session_state.n_revenue = 1

        rev_sources = []
        for ridx in range(st.session_state.n_revenue):
            with st.container():
                st.markdown(f"**収益源 {ridx+1}**")
                rc1, rc2, rc3, rc4 = st.columns(4)
                with rc1:
                    rname = st.text_input("名称", value=f"商品{ridx+1}" if ridx > 0 else "メイン商品", key=f"rname_{ridx}")
                with rc2:
                    rprice = st.number_input(
                        f"平均客単価{annual_label} (円)", value=tmpl["unit_price"] * annual_divisor,
                        step=500, key=f"rprice_{ridx}"
                    )
                    rprice_monthly = rprice / annual_divisor
                with rc3:
                    rweight = st.slider("売上構成比 (%)", 1, 100, 100 if ridx == 0 else 20, key=f"rweight_{ridx}")
                with rc4:
                    rchurn = st.slider("月間解約率 (%)", 0.0, 50.0, tmpl["churn_rate"], 0.5, key=f"rchurn_{ridx}")

                rev_sources.append({
                    "name": rname,
                    "unit_price": rprice_monthly,
                    "weight": rweight / 100,
                    "churn_rate": rchurn / 100,
                })

        rcol1, rcol2 = st.columns(2)
        with rcol1:
            if st.button("＋ 収益源を追加", key="add_rev"):
                st.session_state.n_revenue += 1; st.rerun()
        with rcol2:
            if st.session_state.n_revenue > 1 and st.button("－ 最後の収益源を削除", key="del_rev"):
                st.session_state.n_revenue -= 1; st.rerun()

        st.markdown("---")
        st.markdown('<div class="section-title">集客設計</div>', unsafe_allow_html=True)

        # ─── チャネルタイプ定義 ───
        CHANNEL_TYPES = {
            "リスティング広告": {"icon": "🔍", "color": "#4285f4", "default_cpm": 600,  "default_ctr": 3.0, "default_cvr": 2.5},
            "SNS広告":         {"icon": "📱", "color": "#ea4335", "default_cpm": 400,  "default_ctr": 1.5, "default_cvr": 1.8},
            "ディスプレイ広告": {"icon": "🖥",  "color": "#34a853", "default_cpm": 250,  "default_ctr": 0.5, "default_cvr": 1.2},
            "動画広告":         {"icon": "▶",  "color": "#fbbc04", "default_cpm": 800,  "default_ctr": 1.0, "default_cvr": 1.5},
            "メール/LINE":      {"icon": "📧", "color": "#9b59b6", "default_cpm": 100,  "default_ctr": 5.0, "default_cvr": 3.0},
        }

        # ─── 集客モード切替 ───
        acq_mode_label = st.radio(
            "集客モード",
            ["シンプル (CPA直入力)", "ファネルモード (CTR/CVR)"],
            horizontal=True,
            index=1 if st.session_state.acq_mode == "funnel" else 0,
            key="acq_mode_radio",
        )
        st.session_state.acq_mode = "funnel" if "ファネル" in acq_mode_label else "simple"

        if st.session_state.acq_mode == "simple":
            # ════ シンプルモード ════
            s2a, s2b = st.columns(2)
            with s2a:
                ad_budget = st.number_input(f"月間広告予算{annual_label} (円)", value=tmpl["ad_budget"] * annual_divisor, step=100_000, key="ad_budget_input")
                ad_budget_monthly = ad_budget / annual_divisor
                cpa = st.number_input("CPA (円)", value=tmpl["cpa"], step=100, key="cpa")
            with s2b:
                organic_start = st.number_input("自然流入獲得数 (件/月)", value=tmpl["organic_start"], step=10, key="organic_start")
                org_growth_pct = st.slider("自然流入月次成長率 (%)", 0.0, 20.0, tmpl["organic_growth"], 0.5)
                organic_growth = 1 + org_growth_pct / 100
                k_factor = st.slider("バイラル係数 K factor", 0.0, 0.9, 0.0, 0.05,
                                     help="1ユーザーが紹介する平均新規ユーザー数")
            effective_cpa = cpa
            total_ad_budget_monthly = ad_budget_monthly

        else:
            # ════ ファネルモード ════
            st.markdown("**チャネル追加**")
            ch_btn_cols = st.columns(len(CHANNEL_TYPES))
            for idx, (ch_type, ch_info) in enumerate(CHANNEL_TYPES.items()):
                with ch_btn_cols[idx]:
                    if st.button(f"{ch_info['icon']} {ch_type}", key=f"add_ch_{idx}", use_container_width=True):
                        existing = [ch["name"] for ch in st.session_state.channels]
                        if ch_type not in existing:
                            st.session_state.channels.append({
                                "name": ch_type,
                                "budget": 300_000,
                                "cpm": ch_info["default_cpm"],
                                "ctr": ch_info["default_ctr"],
                                "cvr": ch_info["default_cvr"],
                            })
                            st.rerun()

            # ─── 各チャネルの入力 ───
            channels_data = []
            channels_to_remove = []
            for cidx, ch in enumerate(st.session_state.channels):
                ch_info = CHANNEL_TYPES.get(ch["name"], {"icon": "📢", "color": "#f5a623"})
                c_hdr, c_rm = st.columns([9, 1])
                with c_hdr:
                    st.markdown(f"**{ch_info['icon']} {ch['name']}**")
                with c_rm:
                    if len(st.session_state.channels) > 1 and st.button("✕", key=f"rm_ch_{cidx}"):
                        channels_to_remove.append(cidx)

                cc1, cc2, cc3, cc4 = st.columns(4)
                with cc1:
                    ch_budget = st.number_input("月間予算 (円)", value=int(ch["budget"]), step=50_000, key=f"ch_b_{cidx}")
                with cc2:
                    ch_cpm = st.number_input("CPM (円)", value=float(ch["cpm"]), step=50.0, key=f"ch_cpm_{cidx}", help="1,000インプレッションあたりの単価")
                with cc3:
                    ch_ctr = st.number_input("CTR (%)", value=float(ch["ctr"]), step=0.1, min_value=0.01, key=f"ch_ctr_{cidx}", help="クリック率")
                with cc4:
                    ch_cvr = st.number_input("CVR (%)", value=float(ch["cvr"]), step=0.1, min_value=0.01, key=f"ch_cvr_{cidx}", help="コンバージョン率")

                # 自動計算
                ch_imps = int(ch_budget / ch_cpm * 1000) if ch_cpm > 0 else 0
                ch_clicks = int(ch_imps * ch_ctr / 100)
                ch_cvs = int(ch_clicks * ch_cvr / 100)
                ch_eff_cpa = int(ch_budget / ch_cvs) if ch_cvs > 0 else 0

                # チャネル別ミニファネル
                clr = ch_info.get("color", "#f5a623")
                st.markdown(f"""
                <div class="ch-funnel">
                  <div class="cf-step"><div class="cf-num" style="color:{clr}">¥{ch_budget:,}</div><div class="cf-lbl">予算</div></div>
                  <div class="cf-arr">▶</div>
                  <div class="cf-step"><div class="cf-num" style="color:#8899bb">{ch_imps:,}</div><div class="cf-lbl">imp</div></div>
                  <div class="cf-arr">▶</div>
                  <div class="cf-step"><div class="cf-num" style="color:#8899bb">{ch_clicks:,}</div><div class="cf-lbl">click</div><div class="cf-sub">CTR {ch_ctr}%</div></div>
                  <div class="cf-arr">▶</div>
                  <div class="cf-step"><div class="cf-num" style="color:#34d297">{ch_cvs:,}</div><div class="cf-lbl">CV</div><div class="cf-sub">CVR {ch_cvr}%</div></div>
                  <div class="cf-arr">▶</div>
                  <div class="cf-step"><div class="cf-num" style="color:#fff">¥{ch_eff_cpa:,}</div><div class="cf-lbl">CPA</div></div>
                </div>""", unsafe_allow_html=True)

                channels_data.append({
                    "name": ch["name"], "budget": ch_budget, "cpm": ch_cpm,
                    "ctr": ch_ctr, "cvr": ch_cvr,
                    "impressions": ch_imps, "clicks": ch_clicks, "cvs": ch_cvs, "eff_cpa": ch_eff_cpa,
                })

            # チャネル削除処理
            if channels_to_remove:
                for ci in sorted(channels_to_remove, reverse=True):
                    st.session_state.channels.pop(ci)
                st.rerun()

            # セッション更新
            for cidx, cd in enumerate(channels_data):
                if cidx < len(st.session_state.channels):
                    st.session_state.channels[cidx] = {
                        "name": cd["name"], "budget": cd["budget"],
                        "cpm": cd["cpm"], "ctr": cd["ctr"], "cvr": cd["cvr"],
                    }

            # ─── 合計ファネルサマリー ───
            total_ad_budget_monthly = sum(cd["budget"] for cd in channels_data)
            total_imps  = sum(cd["impressions"] for cd in channels_data)
            total_clicks_all = sum(cd["clicks"] for cd in channels_data)
            total_cvs   = sum(cd["cvs"] for cd in channels_data)
            effective_cpa = int(total_ad_budget_monthly / total_cvs) if total_cvs > 0 else 0
            ad_budget_monthly = total_ad_budget_monthly
            cpa = effective_cpa

            st.markdown(f"""
            <div class="funnel-total">
              <div class="funnel-total-title">▌ チャネル合計ファネル</div>
              <div class="funnel-row">
                <div class="funnel-step"><div class="funnel-num" style="color:#f5a623">¥{total_ad_budget_monthly:,}</div><div class="funnel-lbl">総広告投下</div></div>
                <div class="funnel-arr">▶</div>
                <div class="funnel-step"><div class="funnel-num" style="color:#4285f4">{total_imps:,}</div><div class="funnel-lbl">インプレッション</div></div>
                <div class="funnel-arr">▶</div>
                <div class="funnel-step"><div class="funnel-num" style="color:#9b59b6">{total_clicks_all:,}</div><div class="funnel-lbl">クリック</div></div>
                <div class="funnel-arr">▶</div>
                <div class="funnel-step"><div class="funnel-num" style="color:#34d297">{total_cvs:,}</div><div class="funnel-lbl">CV獲得</div></div>
                <div class="funnel-arr">▶</div>
                <div class="funnel-step"><div class="funnel-num" style="color:#fff">¥{effective_cpa:,}</div><div class="funnel-lbl" style="color:#34d297;font-weight:700;">実質CPA（自動）</div></div>
              </div>
            </div>""", unsafe_allow_html=True)

            # ─── 自然流入 + バイラル ───
            st.markdown("**自然流入 + バイラル設定**")
            s2b1, s2b2, s2b3 = st.columns(3)
            with s2b1:
                organic_start = st.number_input("自然流入獲得数 (件/月)", value=tmpl["organic_start"], step=10)
            with s2b2:
                org_growth_pct = st.slider("自然流入月次成長率 (%)", 0.0, 20.0, tmpl["organic_growth"], 0.5)
                organic_growth = 1 + org_growth_pct / 100
            with s2b3:
                k_factor = st.slider("バイラル係数 K factor", 0.0, 0.9, 0.0, 0.05,
                                     help="1ユーザーが紹介する平均新規ユーザー数 (K=0.3 → 10人→3人紹介)")

        # IMPROVEMENT 5: Year-by-Year Growth Rate Settings
        use_yearly_growth = st.checkbox("年度別成長率を設定", value=False)
        yearly_growth_rates = {}
        if use_yearly_growth:
            st.markdown("**年度別成長率調整（有機成長の倍率）**")
            yr_cols = st.columns(5)
            for year in range(2, 11):
                with yr_cols[(year - 2) % 5]:
                    yearly_growth_rates[year] = st.number_input(
                        f"Year {year}", value=100.0, min_value=0.0, step=10.0, key=f"yr_growth_{year}"
                    ) / 100.0
        else:
            for year in range(2, 11):
                yearly_growth_rates[year] = 1.0

        use_churn = st.checkbox("解約率を反映する", value=True)
        use_season = st.checkbox("季節変動を反映する", value=True)
        if use_season:
            sf_cols = st.columns(6)
            seasonal = []
            for idx, ml in enumerate(MONTH_LABELS):
                with sf_cols[idx % 6]:
                    seasonal.append(st.number_input(ml, 0.1, 3.0, tmpl["seasonal"][idx], 0.05, key=f"sf{idx}"))
        else:
            seasonal = [1.0] * 12

        bc, _, nc = st.columns([1, 7, 1])
        with bc:
            if st.button("戻る", key="b2"):
                st.session_state.step = 1; st.rerun()
        with nc:
            if st.button("次へ", key="n2"):
                st.session_state.step = 3; st.rerun()

    # ═══════════════════════════════════════
    # STEP 3 — コスト設計 (A4: チェックボックス選択式 + IMPROVEMENT 3: More items)
    # ═══════════════════════════════════════
    with st.expander("Step 3 — コスト設計（選択式）", expanded=(st.session_state.step == 3)):
        st.caption("業種テンプレートの項目がデフォルトで選択されています。追加・削除は自由です。")

        s3a, s3b = st.columns(2)

        # --- 変動費 ---
        with s3a:
            st.markdown("**変動費（1件ごと / 売上比率）**")
            vc_values = {}
            for item_name, item_info in VARIABLE_COST_ITEMS.items():
                tmpl_vc = tmpl.get("vc_items", {})
                default_on = item_name in tmpl_vc
                enabled = st.checkbox(
                    f"{item_name}",
                    value=default_on,
                    key=f"vc_chk_{item_info['key']}",
                    help=item_info["desc"],
                )
                if enabled:
                    default_val = tmpl_vc.get(item_name, 0)
                    label = f"  └ {item_name} ({item_info['unit']})"
                    val = st.number_input(
                        label, value=float(default_val), step=0.1 if item_info["unit"] == "%" else 100.0,
                        key=f"vc_val_{item_info['key']}"
                    )
                    vc_values[item_name] = {"value": val, "unit": item_info["unit"]}

            st.markdown("**カスタム変動費の追加**")
            custom_vc_name = st.text_input("カスタム項目名", value="", key="custom_vc_name")
            if custom_vc_name:
                custom_vc_unit = st.radio("単位", ["円/件", "%"], horizontal=True, key="custom_vc_unit")
                custom_vc_val = st.number_input(f"{custom_vc_name} ({custom_vc_unit})", value=0.0, step=100.0 if custom_vc_unit == "円/件" else 0.1, key="custom_vc_val")
                if custom_vc_val > 0:
                    vc_values[custom_vc_name] = {"value": custom_vc_val, "unit": custom_vc_unit}

        # --- 固定費 ---
        with s3b:
            st.markdown("**固定費（月額）**")
            fc_values = {}
            for item_name, item_info in FIXED_COST_ITEMS.items():
                tmpl_fc = tmpl.get("fc_items", {})
                default_on = item_name in tmpl_fc
                enabled = st.checkbox(
                    f"{item_name}",
                    value=default_on,
                    key=f"fc_chk_{item_info['key']}",
                    help=item_info["desc"],
                )
                if enabled:
                    default_val = tmpl_fc.get(item_name, 0)
                    label = f"  └ {item_name}{annual_label} (円)"
                    val = st.number_input(
                        label, value=default_val * annual_divisor, step=10_000,
                        key=f"fc_val_{item_info['key']}"
                    )
                    fc_values[item_name] = val / annual_divisor  # 月額に変換

            st.markdown("**カスタム固定費の追加**")
            custom_fc_name = st.text_input("カスタム項目名", value="", key="custom_fc_name")
            if custom_fc_name:
                custom_fc_val = st.number_input(f"{custom_fc_name} (円/月)", value=0.0, step=10_000, key="custom_fc_val")
                if custom_fc_val > 0:
                    fc_values[custom_fc_name] = custom_fc_val / annual_divisor

        # 固定費合計を計算
        total_fixed = sum(fc_values.values())

        # 変動費の単価・率を整理
        vc_per_unit_fixed = 0  # 円/件ベースの変動費合計
        vc_pct_of_sales = 0    # %ベースの変動費合計
        for item_name, item_data in vc_values.items():
            if item_data["unit"] == "%":
                vc_pct_of_sales += item_data["value"] / 100
            else:
                vc_per_unit_fixed += item_data["value"]

        bc3, _, nc3 = st.columns([1, 7, 1])
        with bc3:
            if st.button("戻る", key="b3"):
                st.session_state.step = 2; st.rerun()
        with nc3:
            if st.button("次へ", key="n3"):
                st.session_state.step = 4; st.rerun()

    # ═══════════════════════════════════════
    # STEP 4 — 人員計画 (B1)
    # ═══════════════════════════════════════
    with st.expander("Step 4 — 人員計画（採用タイミング）", expanded=(st.session_state.step == 4)):
        st.caption("何ヶ月目に何人採用するかを設定すると、人件費がステップ関数で反映されます。")

        if "n_hires" not in st.session_state:
            st.session_state.n_hires = 1

        hire_plan = []
        for hidx in range(st.session_state.n_hires):
            hc1, hc2, hc3, hc4 = st.columns(4)
            with hc1:
                h_month = st.number_input("採用月", min_value=1, max_value=120, value=min(1 + hidx * 6, 120), key=f"hire_month_{hidx}")
            with hc2:
                h_count = st.number_input("人数", min_value=1, max_value=50, value=1, key=f"hire_count_{hidx}")
            with hc3:
                h_salary = st.number_input("月給 (円/人)", value=350_000, step=50_000, key=f"hire_salary_{hidx}")
            with hc4:
                h_role = st.text_input("役職", value="エンジニア" if hidx == 0 else "営業", key=f"hire_role_{hidx}")
            hire_plan.append({"month": h_month, "count": h_count, "salary": h_salary, "role": h_role})

        hcol1, hcol2 = st.columns(2)
        with hcol1:
            if st.button("＋ 採用枠を追加", key="add_hire"):
                st.session_state.n_hires += 1; st.rerun()
        with hcol2:
            if st.session_state.n_hires > 1 and st.button("－ 最後の枠を削除", key="del_hire"):
                st.session_state.n_hires -= 1; st.rerun()

        # 人員計画プレビュー
        if hire_plan:
            st.markdown("**人件費シミュレーション（プレビュー）**")
            preview_months = min(sim_months if 'sim_months' in dir() else 36, 60)
            headcount_data = []
            for m in range(1, preview_months + 1):
                total_heads = 0
                total_salary = 0
                for hp in hire_plan:
                    if m >= hp["month"]:
                        total_heads += hp["count"]
                        total_salary += hp["count"] * hp["salary"]
                # 社会保険料（約15%を自動加算）
                total_cost = total_salary * 1.15
                headcount_data.append({"月": m, "人数": total_heads, "人件費（社保込）": total_cost})
            hc_df = pd.DataFrame(headcount_data)
            st.altair_chart(
                _dk(alt.Chart(hc_df).mark_area(opacity=0.4, color="#f5a623").encode(
                    x=alt.X("月:Q", title="月"),
                    y=alt.Y("人件費（社保込）:Q", axis=alt.Axis(format="~s", title="¥ 月額人件費")),
                    tooltip=["月", "人数", alt.Tooltip("人件費（社保込）:Q", format=",")]
                ).interactive()),
                use_container_width=True
            )

        bc4, _, nc4 = st.columns([1, 7, 1])
        with bc4:
            if st.button("戻る", key="b4"):
                st.session_state.step = 3; st.rerun()
        with nc4:
            if st.button("次へ", key="n4"):
                st.session_state.step = 5; st.rerun()

    # ═══════════════════════════════════════
    # STEP 5 — 設備投資・減価償却 (A3 + IMPROVEMENT 2)
    # ═══════════════════════════════════════
    with st.expander("Step 5 — 設備投資・減価償却", expanded=(st.session_state.step == 5)):
        st.caption("資産を登録すると、定額法または定率法で月割り減価償却費を自動計算し、P&Lとキャッシュフローに反映します。")

        if "n_assets" not in st.session_state:
            st.session_state.n_assets = 0

        dep_assets = []
        for aidx in range(st.session_state.n_assets):
            ac1, ac2, ac3, ac4, ac5 = st.columns(5)
            with ac1:
                a_name = st.text_input("資産名", value=f"資産{aidx+1}", key=f"asset_name_{aidx}")
            with ac2:
                a_cat = st.selectbox("種別", list(DEPRECIATION_CATEGORIES.keys()), key=f"asset_cat_{aidx}")
            with ac3:
                a_cost = st.number_input("取得原価 (円)", value=1_000_000, step=100_000, key=f"asset_cost_{aidx}")
            with ac4:
                default_life = DEPRECIATION_CATEGORIES[a_cat]["useful_life"]
                a_life = st.number_input("耐用年数", min_value=1, max_value=50, value=default_life, key=f"asset_life_{aidx}")
            with ac5:
                a_start = st.number_input("取得月", min_value=0, max_value=120, value=0, key=f"asset_start_{aidx}",
                                          help="0=事業開始前（初期投資）、1=1ヶ月目...")

            # IMPROVEMENT 2: Depreciation method choice
            a_method = st.radio(f"資産{aidx+1} 減価償却方法", ["定額法", "定率法 (200%)"], horizontal=True, key=f"asset_method_{aidx}")
            a_residual = st.number_input(f"資産{aidx+1} 残存価額 (円)", value=1, step=1, key=f"asset_residual_{aidx}", min_value=1)

            if a_method == "定額法":
                monthly_dep = (a_cost - a_residual) / (a_life * 12) if a_life > 0 else 0
            else:  # 定率法
                declining_rate = 2.0 / a_life
                monthly_dep = a_cost * declining_rate / 12

            dep_assets.append({
                "name": a_name, "category": a_cat, "cost": a_cost,
                "useful_life": a_life, "start_month": a_start,
                "method": a_method, "residual_value": a_residual,
                "monthly_dep": monthly_dep,
            })

        acol1, acol2 = st.columns(2)
        with acol1:
            if st.button("＋ 資産を追加", key="add_asset"):
                st.session_state.n_assets += 1; st.rerun()
        with acol2:
            if st.session_state.n_assets > 0 and st.button("－ 最後の資産を削除", key="del_asset"):
                st.session_state.n_assets -= 1; st.rerun()

        if dep_assets:
            st.markdown("**減価償却スケジュール**")
            dep_summary = []
            for a in dep_assets:
                dep_summary.append({
                    "資産名": a["name"],
                    "種別": a["category"],
                    "取得原価": f"¥{a['cost']:,}",
                    "残存価額": f"¥{a['residual_value']:,}",
                    "耐用年数": f"{a['useful_life']}年",
                    "方法": a["method"],
                    "月額償却費": f"¥{a['monthly_dep']:,.0f}",
                    "取得月": f"{a['start_month']}ヶ月目" if a["start_month"] > 0 else "初期投資",
                })
            st.dataframe(pd.DataFrame(dep_summary), hide_index=True, use_container_width=True)

            # IMPROVEMENT 2: Show depreciation schedule chart for each asset
            for a in dep_assets:
                with st.expander(f"減価償却スケジュール詳細 — {a['name']}", expanded=False):
                    dep_schedule = []
                    remaining_value = a['cost']
                    for month in range(1, min(a['useful_life'] * 12 + 1, sim_months + 1)):
                        if a['method'] == "定額法":
                            monthly_depr = (a['cost'] - a['residual_value']) / (a['useful_life'] * 12)
                        else:
                            declining_rate = 2.0 / a['useful_life']
                            monthly_depr = remaining_value * declining_rate / 12
                        remaining_value = max(a['residual_value'], remaining_value - monthly_depr)
                        dep_schedule.append({
                            "月": month,
                            "月額償却費": monthly_depr,
                            "累積償却額": a['cost'] - remaining_value,
                            "帳簿価額": remaining_value,
                        })
                    dep_sch_df = pd.DataFrame(dep_schedule)
                    st.altair_chart(
                        _dk(alt.Chart(dep_sch_df).mark_line().encode(
                            x=alt.X("月:Q", title="月"),
                            y=alt.Y("帳簿価額:Q", axis=alt.Axis(format="~s", title="¥ 帳簿価額")),
                            color=alt.value("#8B5CF6"),
                            tooltip=["月", alt.Tooltip("月額償却費:Q", format=",.0f"), alt.Tooltip("帳簿価額:Q", format=",.0f")]
                        ).interactive()),
                        use_container_width=True
                    )

        bc5, _, nc5 = st.columns([1, 7, 1])
        with bc5:
            if st.button("戻る", key="b5"):
                st.session_state.step = 4; st.rerun()
        with nc5:
            if st.button("次へ", key="n5"):
                st.session_state.step = 6; st.rerun()

    # ═══════════════════════════════════════
    # STEP 6 — 資金繰り
    # ═══════════════════════════════════════
    with st.expander("Step 6 — 資金繰り（キャッシュフロー）", expanded=(st.session_state.step == 6)):
        s6a, s6b = st.columns(2)
        with s6a:
            cash_init = st.number_input("手元現金 (円)", value=10_000_000, step=1_000_000)
            pay_cyc = st.selectbox("入金サイクル", ["当月", "翌月", "翌々月"])
        with s6b:
            exp_cyc = st.selectbox("支払サイクル", ["当月", "翌月"])
            fundraise_alert = st.number_input(
                "資金調達アラート残高 (円)", value=3_000_000, step=500_000,
                help="キャッシュ残高がこの金額を下回ると警告を表示します (B4)"
            )
        pay_delay = {"当月": 0, "翌月": 1, "翌々月": 2}[pay_cyc]
        exp_delay = {"当月": 0, "翌月": 1}[exp_cyc]

        bc6, _, nc6 = st.columns([1, 7, 1])
        with bc6:
            if st.button("戻る", key="b6"):
                st.session_state.step = 5; st.rerun()
        with nc6:
            if st.button("次へ", key="n6"):
                st.session_state.step = 7; st.rerun()

    # ═══════════════════════════════════════
    # STEP 7 — 税効果 (B3)
    # ═══════════════════════════════════════
    with st.expander("Step 7 — 税効果モデル", expanded=(st.session_state.step == 7)):
        st.caption("法人税等の実効税率を設定し、減価償却による税シールド効果を可視化します。")
        tc1, tc2 = st.columns(2)
        with tc1:
            tax_rate_pct = st.slider("実効税率 (%)", 0, 50, 30, help="法人税・住民税・事業税の合計実効税率")
            tax_rate = tax_rate_pct / 100
        with tc2:
            st.info(f"減価償却費は費用として計上され、課税所得を減少させます。\n"
                    f"税シールド効果 = 減価償却費 × 実効税率 ({tax_rate_pct}%)")

        bc7, _ = st.columns([1, 8])
        with bc7:
            if st.button("戻る", key="b7"):
                st.session_state.step = 6; st.rerun()

    # ─── デフォルト補完 ───
    try:
        _ = sim_months
    except NameError:
        sim_months = 36; target_rate = 0.20; target_pct = 20; initial_inv = 5_000_000
        is_annual_input = False; annual_divisor = 1

    try:
        _ = ad_budget_monthly
    except NameError:
        ad_budget_monthly = tmpl["ad_budget"]; cpa = tmpl["cpa"]
        organic_start = tmpl["organic_start"]; organic_growth = 1 + tmpl["organic_growth"] / 100
        use_churn = True; use_season = True; seasonal = tmpl["seasonal"]
        rev_sources = [{"name": "メイン商品", "unit_price": tmpl["unit_price"], "weight": 1.0, "churn_rate": tmpl["churn_rate"] / 100}]

    try:
        _ = effective_cpa
    except NameError:
        effective_cpa = cpa
        total_ad_budget_monthly = ad_budget_monthly

    try:
        _ = k_factor
    except NameError:
        k_factor = 0.0

    try:
        _ = yearly_growth_rates
    except NameError:
        yearly_growth_rates = {year: 1.0 for year in range(2, 11)}

    try:
        _ = total_fixed
    except NameError:
        total_fixed = sum(tmpl.get("fc_items", {}).values())
        vc_per_unit_fixed = sum(v for k, v in tmpl.get("vc_items", {}).items() if k not in ["決済手数料", "モール手数料", "ロイヤリティ"])
        vc_pct_of_sales = sum(v / 100 for k, v in tmpl.get("vc_items", {}).items() if k in ["決済手数料", "モール手数料", "ロイヤリティ"])

    try:
        _ = cash_init
    except NameError:
        cash_init = 10_000_000; pay_delay = 0; exp_delay = 0; fundraise_alert = 3_000_000

    try:
        _ = hire_plan
    except NameError:
        hire_plan = []

    try:
        _ = dep_assets
    except NameError:
        dep_assets = []

    try:
        _ = tax_rate
    except NameError:
        tax_rate = 0.30; tax_rate_pct = 30

    # ─── シナリオ係数 ───
    if scenario == "楽観 +20%":
        s_mult, c_mult = 1.2, 0.9
    elif scenario == "悲観 -20%":
        s_mult, c_mult = 0.8, 1.1
    else:
        s_mult, c_mult = 1.0, 1.0

    # ─── 計算ロジック（全機能統合） ───
    rows = []
    cum_profit = 0
    cum_profit_after_tax = 0
    # 収益源別のアクティブ顧客を管理
    active_by_source = [0] * len(rev_sources)
    cash = cash_init
    s_buf = [0] * (pay_delay + 1)
    e_buf = [0] * (exp_delay + 1)
    bep_m = rec_m = None
    cash_alert_month = None

    # ユニットエコノミクス用の加重平均単価
    weighted_price = sum(rs["unit_price"] * rs["weight"] for rs in rev_sources)
    weighted_churn = sum(rs["churn_rate"] * rs["weight"] for rs in rev_sources) if use_churn else 0

    for i in range(sim_months):
        m = i + 1
        cal = i % 12
        sf = seasonal[cal] if use_season else 1.0

        # IMPROVEMENT 5: Apply yearly growth rate multiplier
        current_year = ((m - 1) // 12) + 1
        year_growth_mult = yearly_growth_rates.get(current_year, 1.0)

        # 集客（全収益源共通の新規獲得）
        u_ad = int(ad_budget_monthly / cpa) if cpa > 0 else 0
        u_org = int(organic_start * (organic_growth ** i) * year_growth_mult)
        # K factor: バイラル係数による増幅
        total_new = int((u_ad + u_org) * (1 + k_factor))

        # 収益源別の売上計算
        total_sales = 0
        total_units = 0
        total_churn = 0
        total_active = 0
        for sidx, rs in enumerate(rev_sources):
            src_new = int(total_new * rs["weight"])
            src_churn = int(active_by_source[sidx] * rs["churn_rate"]) if use_churn and active_by_source[sidx] > 0 else 0
            active_by_source[sidx] = max(0, active_by_source[sidx] + src_new - src_churn)
            src_units = int((active_by_source[sidx] if use_churn else src_new) * sf)
            src_sales = int(src_units * rs["unit_price"] * s_mult)
            total_sales += src_sales
            total_units += src_units
            total_churn += src_churn
            total_active += active_by_source[sidx]

        # 変動費
        vc = (total_units * vc_per_unit_fixed + total_sales * vc_pct_of_sales) * c_mult
        gp = total_sales - vc

        # B1: 人件費ステップ関数
        hire_salary_total = 0
        total_headcount = 0
        for hp in hire_plan:
            if m >= hp["month"]:
                total_headcount += hp["count"]
                hire_salary_total += hp["count"] * hp["salary"]
        hire_cost = hire_salary_total * 1.15  # 社保込み

        # A3: 減価償却費（定額法・定率法対応）
        monthly_depreciation = 0
        for asset in dep_assets:
            if m > asset["start_month"]:
                months_elapsed = m - asset["start_month"]
                total_dep_months = asset["useful_life"] * 12
                if months_elapsed <= total_dep_months:
                    monthly_depreciation += asset["monthly_dep"]

        # 固定費（テンプレ固定費 + 人件費追加分）
        total_fixed_with_hire = total_fixed * c_mult + hire_cost

        # 営業利益（減価償却費を含む）
        op = gp - ad_budget_monthly - total_fixed_with_hire - monthly_depreciation
        cum_profit += op

        # B3: 税効果
        taxable_income = max(0, op)  # 欠損の場合は税ゼロ（簡易モデル）
        tax_amount = taxable_income * tax_rate
        tax_shield = monthly_depreciation * tax_rate  # 減価償却による税シールド
        net_income = op - tax_amount
        cum_profit_after_tax += net_income

        # 損益分岐点
        mr = gp / total_sales if total_sales > 0 else 0
        bep = (total_fixed_with_hire + ad_budget_monthly + monthly_depreciation) / mr if mr > 0 else 0
        if bep_m is None and op > 0:
            bep_m = m
        if rec_m is None and cum_profit > 0:
            rec_m = m

        # CF（減価償却は非現金なのでCFには加算）
        cf_in = total_sales
        cf_out = abs(vc) + ad_budget_monthly + total_fixed_with_hire + tax_amount
        # 設備投資の現金支出
        for asset in dep_assets:
            if m == max(1, asset["start_month"]):
                cf_out += asset["cost"]

        s_buf.append(cf_in)
        e_buf.append(cf_out)
        cash = cash + s_buf.pop(0) - e_buf.pop(0)

        # B4: 資金調達アラート
        if cash_alert_month is None and cash < fundraise_alert:
            cash_alert_month = m

        rows.append({
            "月": f"{m}ヶ月目", "月番号": m, "暦月": MONTH_LABELS[cal],
            "年": ((m - 1) // 12) + 1,
            "季節係数": sf, "新規獲得": total_new, "解約数": total_churn,
            "アクティブ顧客数": total_active, "人員数": total_headcount,
            "販売数": total_units, "売上高": total_sales,
            "変動費": vc, "限界利益": gp,
            "広告宣伝費": ad_budget_monthly,
            "人件費（追加採用）": hire_cost,
            "固定費合計": total_fixed_with_hire,
            "減価償却費": monthly_depreciation,
            "営業利益": op, "累積利益": cum_profit,
            "税額": tax_amount, "税シールド": tax_shield,
            "税引後利益": net_income, "累積税引後利益": cum_profit_after_tax,
            "損益分岐点売上": bep,
            "キャッシュ残高": cash,
            "費用_変動費": vc,
            "費用_広告宣伝費": ad_budget_monthly,
            "費用_固定費": total_fixed_with_hire,
            "費用_減価償却費": monthly_depreciation,
        })

    df = pd.DataFrame(rows)

    # ─── 年次集計 (常に計算：Excel出力・Growth Navigator用) ───
    df_yearly = df.groupby("年").agg({
        "売上高": "sum", "変動費": "sum", "限界利益": "sum",
        "広告宣伝費": "sum", "固定費合計": "sum", "減価償却費": "sum",
        "営業利益": "sum", "累積利益": "last", "累積税引後利益": "last",
        "税額": "sum", "税シールド": "sum", "税引後利益": "sum",
        "損益分岐点売上": "mean", "キャッシュ残高": "last",
        "新規獲得": "sum", "解約数": "sum", "アクティブ顧客数": "last",
        "販売数": "sum", "人員数": "last",
        "費用_変動費": "sum", "費用_広告宣伝費": "sum",
        "費用_固定費": "sum", "費用_減価償却費": "sum",
        "人件費（追加採用）": "sum",
    }).reset_index()

    last = df.iloc[-1]
    cur_sales = last["売上高"]
    cur_profit = last["営業利益"]
    cur_rate = cur_profit / cur_sales if cur_sales > 0 else 0
    gap = cur_sales * target_rate - cur_profit

    # ユニットエコノミクス計算
    ltv = weighted_price / weighted_churn if weighted_churn > 0 else weighted_price * 120
    ltv_cac = ltv / cpa if cpa > 0 else 999

    # ─── B4: 資金調達アラート ───
    if cash_alert_month:
        months_to_alert = cash_alert_month - 1  # 現在=0ヶ月目として
        st.markdown(f"""
        <div class="funding-alert">
            <div class="fa-title">⚠ 資金調達アラート — {cash_alert_month}ヶ月目にキャッシュが ¥{fundraise_alert:,} を下回ります</div>
            <div class="fa-body">
                現在のバーンレートでは <strong>{cash_alert_month}ヶ月目</strong> に資金が不足する見込みです。<br>
                資金調達の準備には通常3〜6ヶ月かかるため、<strong>{max(1, cash_alert_month - 6)}ヶ月目</strong> までに調達活動を開始することを推奨します。<br>
                対策: ① エクイティ調達 ② デットファイナンス ③ コスト削減 ④ 売上加速
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ─── ユニットエコノミクス ───
    st.markdown('<div class="section-title">ユニットエコノミクス</div>', unsafe_allow_html=True)
    margin_per_unit = weighted_price - vc_per_unit_fixed - weighted_price * vc_pct_of_sales
    margin_pct = margin_per_unit / weighted_price * 100 if weighted_price > 0 else 0
    cac_val = cpa
    ltv_val = ltv
    ltv_cac_ratio = ltv_val / cac_val if cac_val > 0 else 999
    payback = cac_val / margin_per_unit if margin_per_unit > 0 else 999
    avg_life = 1 / weighted_churn if weighted_churn > 0 else 120

    ue_html = '<div class="kpi-grid">'
    ue_html += f"""<div class="kpi-card accent">
        <div class="label">客単価 (ARPU)</div><div class="value">¥{weighted_price:,.0f}</div>
        <div class="delta neutral">加重平均（{len(rev_sources)}収益源）</div></div>"""
    ue_html += f"""<div class="kpi-card {'success' if margin_pct > 50 else 'warn'}">
        <div class="label">限界利益 / 件</div><div class="value">¥{margin_per_unit:,.0f}</div>
        <div class="delta {'up' if margin_pct > 50 else 'down'}">利益率 {margin_pct:.1f}%</div></div>"""
    ue_html += f"""<div class="kpi-card accent">
        <div class="label">LTV (顧客生涯価値)</div><div class="value">¥{ltv_val:,.0f}</div>
        <div class="delta neutral">平均 {avg_life:.1f}ヶ月</div></div>"""
    ue_html += f"""<div class="kpi-card {'success' if cac_val < ltv_val / 3 else 'danger'}">
        <div class="label">CAC (獲得単価)</div><div class="value">¥{cac_val:,}</div>
        <div class="delta {'up' if cac_val < ltv_val / 3 else 'down'}">CPA = ¥{cpa:,}</div></div>"""
    ue_html += f"""<div class="kpi-card {'success' if ltv_cac_ratio >= 3 else 'danger'}">
        <div class="label">LTV / CAC</div><div class="value">{ltv_cac_ratio:.1f}x</div>
        <div class="delta {'up' if ltv_cac_ratio >= 3 else 'down'}">{'✓ 健全 (3x以上)' if ltv_cac_ratio >= 3 else '▽ 改善が必要 (3x未満)'}</div></div>"""
    ue_html += f"""<div class="kpi-card {'success' if payback < 12 else 'warn'}">
        <div class="label">ペイバック期間</div><div class="value">{payback:.1f}ヶ月</div>
        <div class="delta {'up' if payback < 12 else 'down'}">{'✓ 12ヶ月以内' if payback < 12 else '▽ 12ヶ月超'}</div></div>"""
    ue_html += "</div>"
    st.markdown(ue_html, unsafe_allow_html=True)

    with st.expander("ユニットエコノミクス詳細を表示"):
        ue_c1, ue_c2 = st.columns(2)
        with ue_c1:
            st.markdown("**収益構造（1顧客あたり）**")
            ue_items = {"売上単価（加重平均）": weighted_price}
            for item_name, item_data in vc_values.items() if 'vc_values' in dir() else []:
                if item_data["unit"] == "%":
                    ue_items[item_name] = -int(weighted_price * item_data["value"] / 100)
                else:
                    ue_items[item_name] = -item_data["value"]
            ue_items["**限界利益**"] = margin_per_unit
            ue_df = pd.DataFrame({"項目": ue_items.keys(), "金額 (円)": [f"¥{v:,.0f}" for v in ue_items.values()]})
            st.dataframe(ue_df, hide_index=True, use_container_width=True)
        with ue_c2:
            st.markdown("**判定基準**")
            checks = [
                ("LTV / CAC ≥ 3.0", ltv_cac_ratio >= 3, f"{ltv_cac_ratio:.1f}x"),
                ("ペイバック ≤ 12ヶ月", payback <= 12, f"{payback:.1f}ヶ月"),
                ("限界利益率 ≥ 50%", margin_pct >= 50, f"{margin_pct:.1f}%"),
                ("解約率 ≤ 5%", weighted_churn * 100 <= 5, f"{weighted_churn*100:.1f}%"),
            ]
            for label, ok, val in checks:
                icon = "✅" if ok else "⚠️"
                st.markdown(f"{icon} **{label}** → 現在: {val}")

    # ─── KPI CARDS (IMPROVEMENT 6: Burn rate metrics) ───
    st.markdown('<div class="section-title">KPI ダッシュボード</div>', unsafe_allow_html=True)

    def kpi(label, value, delta="", delta_type="neutral", accent="accent"):
        return f"""
        <div class="kpi-card {accent}">
            <div class="label">{label}</div>
            <div class="value">{value}</div>
            <div class="delta {delta_type}">{delta}</div>
        </div>"""

    profit_ok = cur_rate >= target_rate
    ltv_ok = ltv_cac >= 3
    run_val = int(cash_init / abs(df["営業利益"].mean())) if df["営業利益"].mean() < 0 else 999
    total_dep_annual = sum(a["monthly_dep"] for a in dep_assets) * 12

    # IMPROVEMENT 6: Calculate burn rate metrics
    last_3_months = df.tail(3)["営業利益"].mean()
    monthly_burn = max(0, -last_3_months) if last_3_months < 0 else 0
    gross_burn = df.tail(1)["費用_変動費"].iloc[0] + df.tail(1)["費用_広告宣伝費"].iloc[0] + df.tail(1)["費用_固定費"].iloc[0] + df.tail(1)["費用_減価償却費"].iloc[0]
    net_burn = gross_burn - df.tail(1)["売上高"].iloc[0]

    kpi_html = '<div class="kpi-grid">'
    kpi_html += kpi("月商", f"¥{cur_sales:,.0f}", f"目標利益率 {target_pct}%", "neutral", "accent")
    kpi_html += kpi("営業利益", f"¥{cur_profit:,.0f}",
                     f"利益率 {cur_rate*100:.1f}% {'✓達成' if profit_ok else '▽未達'}",
                     "up" if profit_ok else "down", "success" if profit_ok else "danger")
    kpi_html += kpi("LTV / CAC", f"{ltv_cac:.1f}x", "3x以上が健全",
                     "up" if ltv_ok else "down", "success" if ltv_ok else "warn")
    kpi_html += kpi("黒字化", f"{bep_m}ヶ月目" if bep_m else "期間外", "", "neutral", "accent")
    kpi_html += kpi("ランウェイ", f"{run_val}ヶ月" if run_val < 999 else "黒字運営", "",
                     "down" if 0 < run_val < 6 else "neutral",
                     "danger" if 0 < run_val < 6 else "accent")
    kpi_html += kpi("チーム人数", f"{int(last['人員数'])} 人",
                     f"追加人件費 ¥{last['人件費（追加採用）']:,.0f}/月", "neutral", "accent")
    kpi_html += kpi("月間バーンレート", f"¥{monthly_burn:,.0f}",
                     "直近3ヶ月平均の営業損失", "down" if monthly_burn > 0 else "neutral",
                     "warn" if monthly_burn > 0 else "success")
    kpi_html += kpi("グロスバーンレート", f"¥{gross_burn:,.0f}",
                     "最終月の総支出", "neutral", "accent")
    kpi_html += kpi("ネットバーンレート", f"¥{net_burn:,.0f}",
                     "支出-収入（負=黒字）", "down" if net_burn > 0 else "up",
                     "danger" if net_burn > 0 else "success")
    kpi_html += "</div>"
    st.markdown(kpi_html, unsafe_allow_html=True)

    # ─── 改善提案 ───
    st.markdown('<div class="section-title">目標達成シミュレーション</div>', unsafe_allow_html=True)
    if profit_ok:
        st.success(f"目標の {target_pct}% を達成しています（現在 {cur_rate*100:.1f}%）")
    else:
        st.warning(f"目標 {target_pct}% まで あと ¥{gap:,.0f}/月 必要です")

    ac1, ac2, ac3, ac4 = st.columns(4)
    pu = gap / last["販売数"] if last["販売数"] > 0 else 0
    with ac1:
        st.markdown(f"""<div class="advice-card">
            <div class="advice-title">単価アップ</div>
            <div class="advice-value">¥{weighted_price+pu:,.0f}</div>
            <div class="advice-desc">+{pu:,.0f}円/件 の値上げ</div></div>""", unsafe_allow_html=True)
    with ac2:
        st.markdown(f"""<div class="advice-card">
            <div class="advice-title">固定費削減</div>
            <div class="advice-value">¥{max(0,total_fixed-gap):,.0f}</div>
            <div class="advice-desc">月額 {gap:,.0f}円 の削減</div></div>""", unsafe_allow_html=True)
    with ac3:
        ncpa = cpa * max(0, 1 - gap / ad_budget_monthly) if ad_budget_monthly > 0 else cpa
        st.markdown(f"""<div class="advice-card">
            <div class="advice-title">CPA 改善</div>
            <div class="advice-value">¥{ncpa:,.0f}</div>
            <div class="advice-desc">現在 ¥{cpa:,} → 目標 ¥{ncpa:,.0f}</div></div>""", unsafe_allow_html=True)
    with ac4:
        nc_rate = weighted_churn * 0.5
        st.markdown(f"""<div class="advice-card">
            <div class="advice-title">解約率半減</div>
            <div class="advice-value">{nc_rate*100:.1f}%</div>
            <div class="advice-desc">現在 {weighted_churn*100:.1f}% → {nc_rate*100:.1f}%</div></div>""", unsafe_allow_html=True)

    # ─── 成長レバー診断 (Growth Navigator v7.0) ───
    st.markdown('<div class="section-title">🧭 成長レバー診断 — Growth Navigator</div>', unsafe_allow_html=True)
    st.caption("あなたの目標を達成するために、どのパラメータを改善すべきか自動診断します。")

    gn_col1, gn_col2 = st.columns([1, 1])
    with gn_col1:
        gn_goal_type = st.selectbox("目標タイプ", [
            "月次黒字化", "年次黒字化（3年以内）", "MAU目標達成",
            "Runway維持（18ヶ月）", "LTV/CAC比 改善"
        ], key="gn_goal_type")
    with gn_col2:
        if gn_goal_type == "MAU目標達成":
            gn_target_val = st.number_input("目標MAU", value=10000, step=1000, key="gn_target_mau")
        elif gn_goal_type == "LTV/CAC比 改善":
            gn_target_val = st.number_input("目標LTV/CAC比", value=3.0, step=0.5, key="gn_target_ltvcac")
        else:
            gn_target_val = None

    # ── ゴール判定 ──
    last_month_profit = last["営業利益"]
    final_mau = int(last["アクティブ顧客数"])
    final_cash = last["キャッシュ残高"]
    monthly_burn = (last["固定費合計"] + last["広告宣伝費"] + last.get("減価償却費", 0))
    runway_months = int(final_cash / monthly_burn) if monthly_burn > 0 else 999

    if gn_goal_type == "月次黒字化":
        bep_month = next((r["月番号"] for _, r in df.iterrows() if r["営業利益"] > 0), None)
        gn_achieved = bep_month is not None
        gn_status_text = f"{bep_month}ヶ月目に達成" if gn_achieved else "期間内に未達"
        gn_gap_desc = "" if gn_achieved else f"最終月の営業損失: ¥{abs(last_month_profit):,.0f}"
    elif gn_goal_type == "年次黒字化（3年以内）":
        yr3_profit = df_yearly[df_yearly["年"] <= 3]["営業利益"].sum() if len(df_yearly[df_yearly["年"] <= 3]) > 0 else 0
        gn_achieved = yr3_profit > 0
        gn_status_text = f"3年累計: ¥{yr3_profit:,.0f}" if gn_achieved else f"3年累計赤字: ¥{yr3_profit:,.0f}"
        gn_gap_desc = "" if gn_achieved else f"黒字化まであと ¥{abs(yr3_profit):,.0f}"
    elif gn_goal_type == "MAU目標達成":
        gn_achieved = final_mau >= gn_target_val
        gn_status_text = f"最終MAU: {final_mau:,}" + (" — 達成!" if gn_achieved else f" / 目標: {int(gn_target_val):,}")
        gn_gap_desc = "" if gn_achieved else f"あと {int(gn_target_val) - final_mau:,} ユーザー不足"
    elif gn_goal_type == "Runway維持（18ヶ月）":
        gn_achieved = runway_months >= 18
        gn_status_text = f"現在のRunway: {runway_months}ヶ月" + (" — 安全圏" if gn_achieved else " — 危険")
        gn_gap_desc = "" if gn_achieved else f"あと {18 - runway_months}ヶ月分の余裕が必要"
    else:  # LTV/CAC
        gn_achieved = ltv_cac >= gn_target_val
        gn_status_text = f"現在のLTV/CAC: {ltv_cac:.2f}" + (" — 健全" if gn_achieved else f" / 目標: {gn_target_val:.1f}")
        gn_gap_desc = "" if gn_achieved else f"LTV/CACを {gn_target_val - ltv_cac:.2f} 改善必要"

    if gn_achieved:
        st.markdown(f"""<div class="gn-goal-card" style="border-color:#4ade80;">
            <div class="gn-title" style="color:#4ade80;">✅ 目標達成</div>
            <div class="gn-value">{gn_status_text}</div></div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"""<div class="gn-goal-card">
            <div class="gn-title">🎯 {gn_goal_type}</div>
            <div class="gn-value">{gn_status_text}</div>
            <div class="gn-sub">{gn_gap_desc}</div></div>""", unsafe_allow_html=True)

    # ── レバー感度分析 ──
    if not gn_achieved:
        st.markdown('<div class="gn-section-header">📊 改善レバー ランキング</div>', unsafe_allow_html=True)
        st.caption("各パラメータを10%改善した場合のインパクトを自動計算。「適用」ボタンでサイドバーの値を即座に更新できます。")

        def quick_kpi(price_mult=1.0, cpa_mult=1.0, churn_mult=1.0, organic_mult=1.0, fc_mult=1.0, ad_mult=1.0):
            """簡易シミュレーション：指定倍率でKPIを再計算"""
            sim_price = weighted_price * price_mult
            sim_cpa = cpa * cpa_mult
            sim_churn = weighted_churn * churn_mult
            sim_organic = organic_start * organic_mult
            sim_fc = total_fixed * fc_mult
            sim_ad = ad_budget_monthly * ad_mult
            # 簡易月次計算
            sim_active = 0
            sim_total_profit = 0
            sim_cash = initial_cash
            for m in range(1, sim_months + 1):
                new_paid = int(sim_ad / sim_cpa) if sim_cpa > 0 else 0
                new_organic = int(sim_organic * (1 + organic_growth / 100) ** ((m - 1) / 12))
                new_total = new_paid + new_organic
                sim_active = max(0, sim_active * (1 - sim_churn) + new_total)
                sales_count = int(sim_active)
                revenue = sales_count * sim_price
                vc = sales_count * (vc_per_unit_fixed + sim_price * vc_pct_of_sales)
                op_profit = revenue - vc - sim_fc - sim_ad
                sim_total_profit += op_profit
                sim_cash += op_profit
            return {"final_profit": op_profit, "total_profit": sim_total_profit,
                    "final_mau": int(sim_active), "final_cash": sim_cash,
                    "ltv_cac": (sim_price / sim_churn if sim_churn > 0 else sim_price * 120) / sim_cpa if sim_cpa > 0 else 999}

        base = quick_kpi()
        levers = [
            {"name": "単価 +10%",       "key": "unit_price",       "params": {"price_mult": 1.1},   "direction": "↑ 単価を上げる"},
            {"name": "CPA -10%",         "key": "cpa",             "params": {"cpa_mult": 0.9},     "direction": "↓ 獲得コストを下げる"},
            {"name": "解約率 -10%",      "key": "churn_rate",      "params": {"churn_mult": 0.9},   "direction": "↓ 解約を減らす"},
            {"name": "オーガニック +10%", "key": "organic_start",   "params": {"organic_mult": 1.1}, "direction": "↑ 自然流入を増やす"},
            {"name": "固定費 -10%",      "key": "fc_total",        "params": {"fc_mult": 0.9},      "direction": "↓ 固定費を削る"},
            {"name": "広告費 +10%",      "key": "ad_budget",       "params": {"ad_mult": 1.1},      "direction": "↑ 広告投資を増やす"},
        ]

        # 各レバーのインパクト計算
        for lv in levers:
            result = quick_kpi(**lv["params"])
            if gn_goal_type in ["月次黒字化", "年次黒字化（3年以内）"]:
                lv["impact"] = result["total_profit"] - base["total_profit"]
                lv["impact_label"] = f"累積利益 {'+' if lv['impact'] >= 0 else ''}{lv['impact']:,.0f}円"
            elif gn_goal_type == "MAU目標達成":
                lv["impact"] = result["final_mau"] - base["final_mau"]
                lv["impact_label"] = f"MAU {'+' if lv['impact'] >= 0 else ''}{lv['impact']:,}"
            elif gn_goal_type == "Runway維持（18ヶ月）":
                lv["impact"] = result["final_cash"] - base["final_cash"]
                lv["impact_label"] = f"キャッシュ {'+' if lv['impact'] >= 0 else ''}¥{lv['impact']:,.0f}"
            else:
                lv["impact"] = result["ltv_cac"] - base["ltv_cac"]
                lv["impact_label"] = f"LTV/CAC {'+' if lv['impact'] >= 0 else ''}{lv['impact']:.2f}"

        levers_sorted = sorted(levers, key=lambda x: abs(x["impact"]), reverse=True)

        for rank, lv in enumerate(levers_sorted, 1):
            impact_color = "#4ade80" if lv["impact"] > 0 else "#f87171" if lv["impact"] < 0 else "#a0b0c8"
            lc1, lc2, lc3 = st.columns([5, 3, 2])
            with lc1:
                st.markdown(f"""<div class="gn-lever-row">
                    <div class="lever-rank">#{rank}</div>
                    <div><div class="lever-name">{lv['name']}</div>
                    <div class="lever-current">{lv['direction']}</div></div>
                </div>""", unsafe_allow_html=True)
            with lc2:
                st.markdown(f'<div style="padding:14px 0;"><span style="color:{impact_color};font-weight:700;font-size:0.9rem;">{lv["impact_label"]}</span></div>', unsafe_allow_html=True)
            with lc3:
                apply_key = f"gn_apply_{lv['key']}"
                if st.button("適用", key=apply_key, help=f"{lv['name']} をサイドバーに反映"):
                    if lv["key"] == "unit_price":
                        for _ri in range(st.session_state.get("n_revenue", 1)):
                            _k = f"rprice_{_ri}"
                            if _k in st.session_state:
                                st.session_state[_k] = int(st.session_state[_k] * 1.1)
                    elif lv["key"] == "cpa":
                        if "cpa" in st.session_state:
                            st.session_state["cpa"] = int(st.session_state["cpa"] * 0.9)
                    elif lv["key"] == "churn_rate":
                        for _ri in range(st.session_state.get("n_revenue", 1)):
                            _k = f"rchurn_{_ri}"
                            if _k in st.session_state:
                                st.session_state[_k] = round(float(st.session_state[_k]) * 0.9, 2)
                    elif lv["key"] == "organic_start":
                        if "organic_start" in st.session_state:
                            st.session_state["organic_start"] = int(st.session_state["organic_start"] * 1.1)
                    elif lv["key"] == "fc_total":
                        for fc_name in selected_fc_items:
                            fc_k = FIXED_COST_ITEMS[fc_name]["key"]
                            if fc_k in st.session_state:
                                st.session_state[fc_k] = int(st.session_state[fc_k] * 0.9)
                    elif lv["key"] == "ad_budget":
                        if "ad_budget_input" in st.session_state:
                            st.session_state["ad_budget_input"] = int(st.session_state["ad_budget_input"] * 1.1)
                    st.rerun()

        # ── 複合効果 ──
        st.markdown('<div class="gn-section-header">🔗 複合効果シミュレーション</div>', unsafe_allow_html=True)
        combo = quick_kpi(price_mult=1.1, cpa_mult=0.9, churn_mult=0.9)
        combo_profit_diff = combo["total_profit"] - base["total_profit"]
        st.markdown(f"""<div class="gn-goal-card">
            <div class="gn-title">単価+10% × CPA-10% × 解約率-10% を同時改善した場合</div>
            <div class="gn-value">累積利益差分: ¥{combo_profit_diff:,.0f}</div>
            <div class="gn-sub">最終月MAU: {combo['final_mau']:,} / 最終キャッシュ: ¥{combo['final_cash']:,.0f}</div>
        </div>""", unsafe_allow_html=True)

    # ── 業種ベンチマーク比較 ──
    st.markdown('<div class="gn-section-header">📈 業種ベンチマーク比較</div>', unsafe_allow_html=True)
    bench = INDUSTRY_BENCHMARKS.get(industry, INDUSTRY_BENCHMARKS["カスタム"])

    actual_churn_pct = weighted_churn * 100
    actual_arpu = weighted_price
    actual_gross_margin = ((cur_sales - last["変動費"]) / cur_sales * 100) if cur_sales > 0 else 0

    bench_items = [
        {"label": "解約率", "actual": actual_churn_pct, "benchmark": bench["churn"], "unit": "%", "lower_better": True},
        {"label": "LTV/CAC比", "actual": ltv_cac, "benchmark": bench["LTV_CAC"], "unit": "x", "lower_better": False},
        {"label": "ARPU", "actual": actual_arpu, "benchmark": bench["ARPU"], "unit": "円", "lower_better": False},
        {"label": "粗利率", "actual": actual_gross_margin, "benchmark": bench["gross_margin"], "unit": "%", "lower_better": False},
    ]

    bc1, bc2 = st.columns(2)
    for idx, bi in enumerate(bench_items):
        col = bc1 if idx % 2 == 0 else bc2
        with col:
            ratio = bi["actual"] / bi["benchmark"] if bi["benchmark"] != 0 else 1
            if bi["lower_better"]:
                bar_color = "#4ade80" if ratio <= 1 else "#f87171"
                status = "良好" if ratio <= 1 else "要改善"
            else:
                bar_color = "#4ade80" if ratio >= 1 else "#f5a623"
                status = "良好" if ratio >= 1 else "要改善"
            fill_w = min(100, ratio * 100) if not bi["lower_better"] else min(100, (1 / ratio) * 100) if ratio > 0 else 100
            st.markdown(f"""<div style="margin:8px 0;">
                <div style="display:flex;justify-content:space-between;margin-bottom:2px;">
                    <span style="color:#c8d8f0;font-size:0.85rem;font-weight:600;">{bi['label']}</span>
                    <span style="color:{bar_color};font-size:0.82rem;font-weight:700;">{status}</span>
                </div>
                <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
                    <span style="color:#8899bb;font-size:0.78rem;">あなた: {bi['actual']:,.1f}{bi['unit']}</span>
                    <span style="color:#a0b0c8;font-size:0.78rem;">業種平均: {bi['benchmark']:,.1f}{bi['unit']}</span>
                </div>
                <div class="gn-benchmark-bar"><div class="gn-benchmark-fill" style="width:{fill_w:.0f}%;background:{bar_color};"></div></div>
            </div>""", unsafe_allow_html=True)

    # ── 競合比較ガイド ──
    with st.expander("🏢 競合に勝つためのシミュレーションガイド"):
        st.markdown("""
        **競合比較の活用方法:**

        1. **解約率が業種平均より高い場合** → サイドバーの解約率を業種平均まで下げてシミュレーション → 改善インパクトを確認
        2. **LTV/CACが低い場合** → 単価アップ or CPA削減を適用 → 目標比に到達するシナリオを探索
        3. **粗利率が低い場合** → 変動費の各項目を見直し → どの費目の削減がもっとも効果的か診断

        **ベンチマーク超えを目指す手順:**
        - 上のレバーランキングで「適用」→ 自動でサイドバー更新 → グラフ・PLが即座に再計算
        - 複合効果で3つ同時改善のシナリオも確認できます
        """)

    # ─── グラフ (IMPROVEMENT 1: Proper legends + IMPROVEMENT 4: Sensitivity analysis) ───
    st.markdown('<div class="section-title">グラフ分析</div>', unsafe_allow_html=True)

    graph_col1, graph_col2 = st.columns([3, 1])
    with graph_col2:
        default_view = "年単位" if sim_months >= 60 else "月単位"
        view_mode = st.radio("表示単位", ["月単位", "年単位"], horizontal=True, key="view_mode",
                             index=0 if default_view == "月単位" else 1)

    if view_mode == "年単位":
        df_yearly["月"] = df_yearly["年"].apply(lambda y: f"{int(y)}年目")
        df_yearly["月番号"] = df_yearly["年"]
        df_view = df_yearly
        x_field = "月番号:Q"
        x_title = "年"
    else:
        df_view = df
        x_field = "月番号:Q"
        x_title = "月"

    # IMPROVEMENT 1 & 4: グラフタブに感度分析を追加
    g1, g2, g3, g4, g5, g6, g7, g8, g9 = st.tabs([
        "収支推移", "キャッシュフロー", "シナリオ比較",
        "コスト構造", "顧客推移", "税効果",
        "LTV/CAC推移", "感度分析", "データ表",
    ])

    with g1:
        # IMPROVEMENT 1: 凡例付き収支推移チャート
        sales_line = alt.Chart(df_view).mark_line(strokeWidth=2.5).encode(
            x=alt.X(x_field, title=x_title),
            y=alt.Y("売上高:Q", axis=alt.Axis(format="~s", title="金額 (¥)")),
            color=alt.value("#2196F3"),
            tooltip=["月", alt.Tooltip("売上高:Q", format=",")]
        )
        bep_line = alt.Chart(df_view).mark_line(strokeDash=[4, 4]).encode(
            x=alt.X(x_field),
            y=alt.Y("損益分岐点売上:Q"),
            color=alt.value("#94A3B8"),
        )
        profit_area = alt.Chart(df_view).mark_area(opacity=0.25).encode(
            x=alt.X(x_field),
            y=alt.Y("営業利益:Q"),
            color=alt.condition(alt.datum.営業利益 > 0, alt.value("#16A34A"), alt.value("#DC2626")),
        )
        # 凡例データ
        legend_df = pd.DataFrame([
            {"label": "売上高", "y": 0, "x": 0},
            {"label": "損益分岐点", "y": 0, "x": 0},
            {"label": "営業利益", "y": 0, "x": 0},
        ])
        legend_chart = alt.Chart(legend_df).mark_point(size=0).encode(
            color=alt.Color("label:N",
                scale=alt.Scale(domain=["売上高", "損益分岐点", "営業利益"],
                                range=["#2196F3", "#94A3B8", "#16A34A"]),
                legend=alt.Legend(title="凡例", orient="top"))
        )
        st.altair_chart(_dk((profit_area + sales_line + bep_line + legend_chart).interactive()), use_container_width=True)

    with g2:
        # IMPROVEMENT 1: キャッシュフロー with 凡例 + B4: アラートライン
        cf_chart = alt.Chart(df_view).mark_area(opacity=0.5).encode(
            x=alt.X(x_field, title=x_title),
            y=alt.Y("キャッシュ残高:Q", axis=alt.Axis(format="~s", title="¥ キャッシュ残高")),
            color=alt.condition(alt.datum.キャッシュ残高 > 0, alt.value("#2196F3"), alt.value("#DC2626")),
            tooltip=["月", alt.Tooltip("キャッシュ残高:Q", format=",")]
        )
        zero_line = alt.Chart(pd.DataFrame({"y": [0]})).mark_rule(color="#DC2626", strokeDash=[3, 3]).encode(y="y:Q")
        alert_line = alt.Chart(pd.DataFrame({"y": [fundraise_alert if 'fundraise_alert' in dir() else 3_000_000]})).mark_rule(
            color="#F59E0B", strokeDash=[6, 3], strokeWidth=2
        ).encode(y="y:Q")

        alert_label = alt.Chart(pd.DataFrame({"y": [fundraise_alert if 'fundraise_alert' in dir() else 3_000_000], "label": ["調達アラートライン"]})).mark_text(
            align="left", dx=5, dy=-8, fontSize=11, color="#F59E0B", fontWeight="bold"
        ).encode(y="y:Q", text="label:N")

        st.altair_chart(_dk((cf_chart + zero_line + alert_line + alert_label).interactive()), use_container_width=True)

    with g3:
        def sc_calc(sm, cm, label):
            r = []; cum = 0; ac_list = [0] * len(rev_sources)
            for i in range(sim_months):
                u_ad2 = int(ad_budget_monthly / cpa) if cpa > 0 else 0
                nc2 = u_ad2 + int(organic_start * (organic_growth ** i))
                total_s = 0
                for sidx, rs in enumerate(rev_sources):
                    sn = int(nc2 * rs["weight"])
                    sch = int(ac_list[sidx] * rs["churn_rate"]) if use_churn else 0
                    ac_list[sidx] = max(0, ac_list[sidx] + sn - sch)
                    su = int((ac_list[sidx] if use_churn else sn) * (seasonal[i % 12] if use_season else 1))
                    total_s += int(su * rs["unit_price"] * sm)
                vc2 = (total_s * vc_pct_of_sales + int(nc2 * (seasonal[i % 12] if use_season else 1)) * vc_per_unit_fixed) * cm
                hc2 = sum(hp["count"] * hp["salary"] * 1.15 for hp in hire_plan if (i + 1) >= hp["month"])
                dep2 = sum(a["monthly_dep"] for a in dep_assets if (i + 1) > a["start_month"] and (i + 1 - a["start_month"]) <= a["useful_life"] * 12)
                cum += total_s - vc2 - ad_budget_monthly - total_fixed * cm - hc2 - dep2
                r.append({"月番号": i + 1, "累積利益": cum, "シナリオ": label})
            return pd.DataFrame(r)

        df_all = pd.concat([sc_calc(1.2, 0.9, "楽観"), sc_calc(1.0, 1.0, "中庸"), sc_calc(0.8, 1.1, "悲観")])
        sc_ch = alt.Chart(df_all).mark_line(strokeWidth=2).encode(
            x=alt.X("月番号:Q", title="月"),
            y=alt.Y("累積利益:Q", axis=alt.Axis(format="~s", title="¥ 累積利益")),
            color=alt.Color("シナリオ:N",
                scale=alt.Scale(domain=["楽観", "中庸", "悲観"],
                                range=["#16A34A", "#2196F3", "#DC2626"]),
                legend=alt.Legend(title="シナリオ", orient="top")),
            tooltip=["月番号", "シナリオ", alt.Tooltip("累積利益:Q", format=",")]
        )
        st.altair_chart(_dk(sc_ch.interactive()), use_container_width=True)

    with g4:
        # IMPROVEMENT 1: コスト構造 with 明示凡例
        cost_cols = ["費用_変動費", "費用_広告宣伝費", "費用_固定費", "費用_減価償却費"]
        cost_labels = ["変動費", "広告宣伝費", "固定費（人件費込）", "減価償却費"]
        cost_colors = ["#F59E0B", "#EF4444", "#6366F1", "#8B5CF6"]

        cd = df_view.melt(id_vars=["月"], value_vars=cost_cols, var_name="費用種別", value_name="金額")
        cd["費用種別"] = cd["費用種別"].map(dict(zip(cost_cols, cost_labels)))

        st.altair_chart(
            _dk(alt.Chart(cd).mark_area().encode(
                x=alt.X("月:N", sort=None, title=x_title),
                y=alt.Y("金額:Q", axis=alt.Axis(format="~s", title="¥ 金額")),
                color=alt.Color("費用種別:N",
                    scale=alt.Scale(domain=cost_labels, range=cost_colors),
                    legend=alt.Legend(title="費用区分", orient="top")),
                tooltip=["月", "費用種別", alt.Tooltip("金額:Q", format=",")]
            )),
            use_container_width=True
        )

    with g5:
        if use_churn:
            # IMPROVEMENT 1: 顧客推移 with 凡例
            cust_base = alt.Chart(df_view).encode(x=alt.X(x_field, title=x_title))
            new_bar = cust_base.mark_bar(opacity=0.6).encode(
                y=alt.Y("新規獲得:Q", axis=alt.Axis(format=",", title="人数")),
                color=alt.value("#BBF7D0"),
            )
            active_line = cust_base.mark_line(strokeWidth=2.5).encode(
                y=alt.Y("アクティブ顧客数:Q"),
                color=alt.value("#2196F3"),
            )
            legend_cust = alt.Chart(pd.DataFrame([
                {"label": "新規獲得", "y": 0}, {"label": "アクティブ顧客数", "y": 0}
            ])).mark_point(size=0).encode(
                color=alt.Color("label:N",
                    scale=alt.Scale(domain=["新規獲得", "アクティブ顧客数"],
                                    range=["#BBF7D0", "#2196F3"]),
                    legend=alt.Legend(title="凡例", orient="top"))
            )
            st.altair_chart(_dk((new_bar + active_line + legend_cust).interactive()), use_container_width=True)
        else:
            st.info("解約率をONにすると顧客推移グラフが表示されます")

    with g6:
        # B3: 税効果チャート
        tax_base = alt.Chart(df_view).encode(x=alt.X(x_field, title=x_title))
        op_line = tax_base.mark_line(strokeWidth=2).encode(
            y=alt.Y("営業利益:Q", axis=alt.Axis(format="~s", title="金額 (¥)")),
            color=alt.value("#2196F3"),
        )
        net_line = tax_base.mark_line(strokeWidth=2).encode(
            y=alt.Y("税引後利益:Q"),
            color=alt.value("#16A34A"),
        )
        tax_bar = tax_base.mark_bar(opacity=0.3).encode(
            y=alt.Y("税額:Q"),
            color=alt.value("#EF4444"),
        )
        shield_line = tax_base.mark_line(strokeDash=[4, 4], strokeWidth=1.5).encode(
            y=alt.Y("税シールド:Q"),
            color=alt.value("#8B5CF6"),
        )
        legend_tax = alt.Chart(pd.DataFrame([
            {"label": "営業利益", "y": 0}, {"label": "税引後利益", "y": 0},
            {"label": "税額", "y": 0}, {"label": "税シールド（償却）", "y": 0},
        ])).mark_point(size=0).encode(
            color=alt.Color("label:N",
                scale=alt.Scale(
                    domain=["営業利益", "税引後利益", "税額", "税シールド（償却）"],
                    range=["#2196F3", "#16A34A", "#EF4444", "#8B5CF6"]),
                legend=alt.Legend(title="凡例", orient="top"))
        )
        st.altair_chart(_dk((tax_bar + op_line + net_line + shield_line + legend_tax).interactive()), use_container_width=True)

    with g7:
        # B5: LTV/CAC推移チャート
        ltv_cac_data = []
        ac_tracking = [0] * len(rev_sources)
        for i in range(sim_months):
            m = i + 1
            sf = seasonal[i % 12] if use_season else 1.0
            u_ad2 = int(ad_budget_monthly / cpa) if cpa > 0 else 0
            nc2 = u_ad2 + int(organic_start * (organic_growth ** i))
            total_s = 0; total_u = 0; w_churn = 0
            for sidx, rs in enumerate(rev_sources):
                sn = int(nc2 * rs["weight"])
                sch = int(ac_tracking[sidx] * rs["churn_rate"]) if use_churn else 0
                ac_tracking[sidx] = max(0, ac_tracking[sidx] + sn - sch)
                su = int((ac_tracking[sidx] if use_churn else sn) * sf)
                total_s += int(su * rs["unit_price"])
                total_u += su
                w_churn += rs["churn_rate"] * rs["weight"]

            if total_u > 0 and w_churn > 0:
                arpu = total_s / total_u
                m_ltv = arpu / w_churn
                m_cac = cpa
                m_ratio = m_ltv / m_cac if m_cac > 0 else 0
            else:
                m_ltv = 0; m_cac = cpa; m_ratio = 0

            ltv_cac_data.append({"月番号": m, "LTV": m_ltv, "CAC": m_cac, "LTV/CAC比率": m_ratio})

        lc_df = pd.DataFrame(ltv_cac_data)
        # LTV/CAC比率の推移
        ratio_line = alt.Chart(lc_df).mark_line(strokeWidth=2.5, color="#2196F3").encode(
            x=alt.X("月番号:Q", title="月"),
            y=alt.Y("LTV/CAC比率:Q", title="LTV / CAC 比率"),
            tooltip=["月番号", alt.Tooltip("LTV/CAC比率:Q", format=".1f"),
                      alt.Tooltip("LTV:Q", format=",.0f"), alt.Tooltip("CAC:Q", format=",")]
        )
        health_line = alt.Chart(pd.DataFrame({"y": [3.0], "label": ["健全ライン (3.0x)"]})).mark_rule(
            color="#16A34A", strokeDash=[6, 3], strokeWidth=2
        ).encode(y="y:Q")
        health_text = alt.Chart(pd.DataFrame({"y": [3.0], "label": ["3.0x 健全ライン"]})).mark_text(
            align="left", dx=5, dy=-8, fontSize=11, color="#16A34A", fontWeight="bold"
        ).encode(y="y:Q", text="label:N")

        st.altair_chart(_dk((ratio_line + health_line + health_text).interactive()), use_container_width=True)

        # LTV vs CAC 金額推移
        lc_melt = lc_df.melt(id_vars=["月番号"], value_vars=["LTV", "CAC"], var_name="指標", value_name="金額")
        lc_chart2 = alt.Chart(lc_melt).mark_line(strokeWidth=2).encode(
            x=alt.X("月番号:Q", title="月"),
            y=alt.Y("金額:Q", axis=alt.Axis(format="~s", title="¥ 金額")),
            color=alt.Color("指標:N",
                scale=alt.Scale(domain=["LTV", "CAC"], range=["#2196F3", "#EF4444"]),
                legend=alt.Legend(title="指標", orient="top")),
            tooltip=["月番号", "指標", alt.Tooltip("金額:Q", format=",")]
        )
        st.altair_chart(_dk(lc_chart2.interactive()), use_container_width=True)

    with g8:
        # IMPROVEMENT 4: 感度分析 (トルネードチャート)
        st.markdown("**パラメータの±20%変動がファイナル月の営業利益に与える影響**")

        sensitivity_params = {
            "客単価": ("weighted_price", 0.8, 1.2),
            "CPA": ("cpa", 0.8, 1.2),
            "広告予算": ("ad_budget_monthly", 0.8, 1.2),
            "固定費": ("total_fixed", 0.8, 1.2),
            "変動費率": ("vc_pct_of_sales", 0.8, 1.2),
            "解約率": ("weighted_churn", 0.8, 1.2),
        }

        sensitivity_results = []
        baseline_profit = last["営業利益"]

        for param_name, (param_var, low_mult, high_mult) in sensitivity_params.items():
            # シミュレーション用のベース値を取得
            base_val = eval(param_var) if param_var in ["weighted_price", "cpa", "ad_budget_monthly", "total_fixed", "vc_pct_of_sales", "weighted_churn"] else 0

            # 低シナリオと高シナリオを計算
            for scenario_mult, scenario_name in [(low_mult, "Low"), (high_mult, "High")]:
                temp_profit = baseline_profit
                if param_var == "weighted_price":
                    temp_price = weighted_price * scenario_mult
                    temp_margin = temp_price - vc_per_unit_fixed - temp_price * vc_pct_of_sales
                    temp_profit = (last["販売数"] * temp_margin) - ad_budget_monthly - total_fixed_with_hire - monthly_depreciation
                elif param_var == "cpa":
                    temp_cpa = cpa * scenario_mult
                    temp_new = int(ad_budget_monthly / temp_cpa) if temp_cpa > 0 else 0
                    temp_profit = last["売上高"] - last["変動費"] - ad_budget_monthly - total_fixed_with_hire - monthly_depreciation
                elif param_var == "ad_budget_monthly":
                    temp_ad = ad_budget_monthly * scenario_mult
                    temp_profit = last["売上高"] - last["変動費"] - temp_ad - total_fixed_with_hire - monthly_depreciation
                elif param_var == "total_fixed":
                    temp_fixed = total_fixed * scenario_mult
                    temp_profit = last["売上高"] - last["変動費"] - ad_budget_monthly - (temp_fixed + hire_cost) - monthly_depreciation
                elif param_var == "vc_pct_of_sales":
                    temp_vc_pct = vc_pct_of_sales * scenario_mult
                    temp_vc = last["販売数"] * vc_per_unit_fixed + last["売上高"] * temp_vc_pct
                    temp_profit = last["売上高"] - temp_vc - ad_budget_monthly - total_fixed_with_hire - monthly_depreciation
                elif param_var == "weighted_churn":
                    temp_churn = weighted_churn * scenario_mult
                    temp_profit = baseline_profit

                sensitivity_results.append({
                    "parameter": param_name,
                    "scenario": scenario_name,
                    "profit_change": temp_profit - baseline_profit
                })

        sens_df = pd.DataFrame(sensitivity_results)
        sens_pivot = sens_df.pivot(index="parameter", columns="scenario", values="profit_change").reset_index()
        sens_pivot["Impact Range"] = sens_pivot["High"] - sens_pivot["Low"]
        sens_pivot = sens_pivot.sort_values("Impact Range", ascending=True)

        # トルネードチャートの描画
        tornado_data = []
        for idx, row in sens_pivot.iterrows():
            tornado_data.append({"Parameter": row["parameter"], "Impact": row["Low"], "Direction": "Low"})
            tornado_data.append({"Parameter": row["parameter"], "Impact": row["High"], "Direction": "High"})

        tornado_df = pd.DataFrame(tornado_data)
        tornado_chart = alt.Chart(tornado_df).mark_bar().encode(
            x=alt.X("Impact:Q", title="営業利益への影響 (¥)"),
            y=alt.Y("Parameter:N", title="パラメータ", sort=list(sens_pivot["parameter"])),
            color=alt.Color("Direction:N",
                scale=alt.Scale(domain=["Low", "High"], range=["#EF4444", "#16A34A"]),
                legend=alt.Legend(title="変動方向", orient="right")),
            tooltip=["Parameter", alt.Tooltip("Impact:Q", format=",.0f"), "Direction"]
        )
        st.altair_chart(_dk(tornado_chart), use_container_width=True)

    with g9:
        st.dataframe(df_view, use_container_width=True)

    # ─── エクスポート ───
    st.markdown('<div class="section-title">データエクスポート</div>', unsafe_allow_html=True)
    ec1, ec2, ec3 = st.columns(3)
    with ec1:
        st.download_button("CSV ダウンロード", df.to_csv(index=False).encode("utf-8-sig"),
                           "simulation.csv", "text/csv", use_container_width=True)
    with ec2:
        try:
            excel_bytes = build_excel_pl(
                df, df_yearly,
                rev_sources=[{"name": s, "share": 1.0 / max(1, len(revenue_sources))} for s in revenue_sources],
                fc_values_dict={n: fc_values.get(n, 0) for n in selected_fc_items} if 'selected_fc_items' in dir() else {},
                vc_values_dict={n: vc_values.get(n, 0) for n in selected_vc_items} if 'selected_vc_items' in dir() else {},
                industry_name=industry,
                company_name=industry,
            )
            st.download_button("Excel P&L ダウンロード", excel_bytes, "BizMaker_PL.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        except Exception as e:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                df.to_excel(w, index=False, sheet_name="PL")
            st.download_button("Excel ダウンロード (簡易版)", buf.getvalue(), "simulation.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with ec3:
        st.button("PDF レポート（Phase 2）", disabled=True, use_container_width=True)


# ═══════════════════════════════════════════════
# TAB 2 — AI アドバイザー（UI モック）
# ═══════════════════════════════════════════════
with tab_ai:
    st.markdown("""
    <div class="cs-banner">
        ⚡ <strong>Coming Soon — Phase 2</strong>
        このタブは機能イメージです。ヒアリング用プレビューとしてご確認ください。
    </div>
    """, unsafe_allow_html=True)

    st.markdown("#### AI アドバイザー")
    st.caption("あなたの事業計画をリアルタイムで分析し、具体的な改善提案を自動生成します。")

    st.markdown("""
    <div style="max-width:680px; margin-top:8px;">
        <div class="ai-bubble">
            <div class="ai-label">Biz Maker AI · 分析結果</div>
            入力された事業計画を分析しました。以下が主な所見です。<br><br>
            <strong>1. キャッシュフロー警告</strong><br>
            現在の入金サイクルと支払サイクルのズレにより、4〜6ヶ月目にキャッシュがタイトになる可能性があります。
            運転資金として 300〜500万円の予備を確保することを推奨します。<br><br>
            <strong>2. LTV / CAC 比率</strong><br>
            現在の比率は 2.4x で、業界健全ラインの 3.0x を下回っています。
            解約率を 1〜2% 改善するだけで比率が 3.6x まで改善し、収益性が大幅に向上します。<br><br>
            <strong>3. 季節変動リスク</strong><br>
            SaaS業種の場合、7〜8月に売上が約 5% 低下する傾向があります。
            この時期に合わせた年次契約プランの提供を検討してみてください。
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**AI に質問する**")
    col_q, col_send = st.columns([6, 1])
    with col_q:
        user_q = st.text_input("", placeholder="例: 解約率を改善するための具体的な施策を教えて", label_visibility="collapsed")
    with col_send:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("送信", use_container_width=True):
            st.info("Phase 2 では Claude API を接続し、リアルタイムで回答します。")

    st.markdown('<div class="section-title">よくある質問</div>', unsafe_allow_html=True)
    qa_cols = st.columns(3)
    questions = [
        ("解約率の改善策", "カスタマーサクセスの強化とオンボーディング改善が最も効果的です。"),
        ("資金調達のタイミング", "ランウェイが6ヶ月を下回る前に調達活動を始めることを推奨します。"),
        ("CPA を下げる方法", "SEO強化による自然流入の増加と、リターゲティング広告の最適化が有効です。"),
    ]
    for col, (q, a) in zip(qa_cols, questions):
        with col:
            with st.expander(q):
                st.markdown(f"<div style='font-size:0.83rem;color:#c8d8f0;line-height:1.6;'>{a}<br><br><em style='color:#8899bb;'>Phase 2 では AIが事業計画データを参照した上で回答します。</em></div>", unsafe_allow_html=True)

    st.markdown('<div class="section-title">会話例プレビュー</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="chat-wrap" style="background:#F8FAFC;border-radius:12px;padding:16px;border:1px solid #E8ECF0;">
        <div style="float:right;clear:both;">
            <div class="user-bubble">解約率を下げるにはどうすればいいですか？</div>
        </div>
        <div style="clear:both; margin-top:8px;">
            <div class="ai-bubble" style="max-width:90%;">
                <div class="ai-label">Biz Maker AI</div>
                解約率改善には主に3つのアプローチが効果的です：<br>
                ① オンボーディングの強化（最初の30日が鍵）<br>
                ② プロダクト内での価値提供の可視化（ダッシュボード等）<br>
                ③ ヘルススコアによる早期チャーン予測と介入<br><br>
                御社の現在の解約率を改善した場合、LTVが大幅に向上し、収益性が改善します。
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════
# TAB 3 — 専門家に相談（UI モック）
# ═══════════════════════════════════════════════
with tab_cons:
    st.markdown("""
    <div class="cs-banner">
        ⚡ <strong>Coming Soon — Phase 2</strong>
        このタブは機能イメージです。ヒアリング用プレビューとしてご確認ください。
    </div>
    """, unsafe_allow_html=True)

    st.markdown("#### 専門家マッチング")
    st.caption("あなたの事業フェーズと課題に合った専門家を見つけて、直接相談できます。")

    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        st.selectbox("専門分野", ["すべて","財務・会計","マーケティング","法務","IT・開発","人事・組織"])
    with fc2:
        st.selectbox("料金帯", ["すべて","〜5,000円/30分","5,000〜10,000円","10,000円〜"])
    with fc3:
        st.selectbox("評価", ["すべて","★ 4.5以上","★ 4.0以上"])
    with fc4:
        st.selectbox("対応形式", ["すべて","オンライン","対面","電話"])

    st.markdown("---")
    consultants = [
        {"initials":"TT","name":"田中 太郎","field":"財務・会計",
         "desc":"公認会計士。スタートアップの資金調達・事業計画策定を 200社以上支援。元BigFour出身。SaaSビジネスの財務モデル設計が専門。",
         "rating":"4.9","reviews":128,"price":"¥8,000 / 30分","badge":"あなたの事業計画にマッチ","tags":["財務モデル","資金調達","SaaS"]},
        {"initials":"HK","name":"鈴木 花子","field":"マーケティング",
         "desc":"元Google。D2C・SaaSのグロースマーケティングを専門とし、CPA改善・LTV向上の実績多数。コンテンツSEOからPaid Socialまで幅広く対応。",
         "rating":"4.8","reviews":94,"price":"¥10,000 / 30分","badge":"CPA改善の実績多数","tags":["グロース","SEO","広告運用"]},
        {"initials":"IY","name":"山田 一郎","field":"法務",
         "desc":"弁護士。スタートアップの法務全般（利用規約・プライバシーポリシー・契約書作成）からIPO準備まで一気通貫で対応。初回30分無料。",
         "rating":"4.6","reviews":67,"price":"¥12,000 / 30分","badge":"初回無料相談あり","tags":["契約書","IPO","規約作成"]},
    ]
    for cons in consultants:
        st.markdown(f"""
        <div class="consultant-card">
            <div style="display:flex; gap:16px; align-items:flex-start;">
                <div style="width:52px;height:52px;border-radius:50%;background:linear-gradient(135deg,#667EEA,#764BA2);
                    display:flex;align-items:center;justify-content:center;color:white;font-weight:700;font-size:1.1rem;flex-shrink:0;">
                    {cons['initials']}
                </div>
                <div style="flex:1;">
                    <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;">
                        <span class="cons-name">{cons['name']}</span>
                        <span class="cons-field">{cons['field']}</span>
                        <span class="cons-badge">{cons['badge']}</span>
                    </div>
                    <div class="cons-desc">{cons['desc']}</div>
                    <div style="display:flex;gap:16px;align-items:center;flex-wrap:wrap;">
                        <div class="cons-meta">★ {cons['rating']} ({cons['reviews']} 件のレビュー)</div>
                        <div class="cons-meta" style="font-weight:600;color:#c8d8f0;">{cons['price']}</div>
                        <div>{''.join(f'<span class="tag">{t}</span>' for t in cons['tags'])}</div>
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button(f"相談を予約する — {cons['name']}", key=f"book_{cons['name']}"):
            st.info("Phase 2 では専門家のカレンダーと連携して直接予約できます。")


# ═══════════════════════════════════════════════
# TAB 4 — コミュニティ（UI モック）
# ═══════════════════════════════════════════════
with tab_sns:
    st.markdown("""
    <div class="cs-banner">
        ⚡ <strong>Coming Soon — Phase 2</strong>
        このタブは機能イメージです。ヒアリング用プレビューとしてご確認ください。
    </div>
    """, unsafe_allow_html=True)

    st.markdown("#### コミュニティ")
    st.caption("同じフェーズの起業家・経営者と繋がり、事業計画のフィードバックを交換しましょう。")

    with st.expander("投稿を作成する", expanded=False):
        post_txt = st.text_area("内容", placeholder="事業計画について相談したいこと、学んだことをシェアしましょう…", height=100)
        tag_opts = st.multiselect("タグ", ["#SaaS","#EC","#飲食","#コンサル","#資金調達","#マーケ","#解約率改善","#初期顧客獲得"])
        if st.button("投稿する", key="post_btn"):
            st.info("Phase 2 で実装予定です。")

    st.markdown("---")
    posts = [
        {"initials":"SM","name":"佐藤 美咲","sub":"SaaS · 創業2年目",
         "content":"解約率を 8% → 2.8% に改善できました\n施策は「オンボーディングフロー」の抜本的な見直し。特に初回ログインから7日間のメール自動化が効きました。",
         "tags":["#SaaS","#解約率改善"],"likes":38,"comments":14,"time":"2時間前"},
        {"initials":"KT","name":"高橋 健太","sub":"飲食店 · 2店舗運営",
         "content":"このシミュレーターで12月の季節変動を入れてシミュレーションしてみたら、固定費の比率が高すぎることに気づきました。業務委託の比率を見直して月30万円のコスト改善ができそうです。",
         "tags":["#飲食","#固定費削減"],"likes":22,"comments":9,"time":"5時間前"},
        {"initials":"RW","name":"渡辺 翔","sub":"コンサルティング · 独立1年目",
         "content":"LTV/CAC が 2.1x で悩んでいます。コンサルビジネスでCACを下げた方法を教えてください。今は紹介プログラムの導入を検討中です。",
         "tags":["#コンサル","#LTV","#CAC"],"likes":45,"comments":21,"time":"1日前"},
    ]
    for post in posts:
        content_html = post["content"].replace("\n", "<br>")
        tags_html = " ".join(f'<span class="tag">{t}</span>' for t in post["tags"])
        st.markdown(f"""
        <div class="post-card">
            <div class="post-header">
                <div class="post-avatar">{post['initials']}</div>
                <div class="post-meta">
                    <div class="name">{post['name']}</div>
                    <div class="sub">{post['sub']} · {post['time']}</div>
                </div>
            </div>
            <div class="post-content">{content_html}</div>
            <div style="margin-top:8px;">{tags_html}</div>
            <div class="post-actions">
                <span class="post-action">👍 {post['likes']}</span>
                <span class="post-action">💬 {post['comments']} コメント</span>
                <span class="post-action">🔖 保存</span>
                <span class="post-action">↗ シェア</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button(f"コメントする — {post['name']}", key=f"cmt_{post['name']}", type="secondary"):
            st.info("Phase 2 で実装予定です。")

# ─── フッター ───
st.markdown("""
<div style="margin-top:3rem;padding-top:1rem;border-top:1px solid #1c2b44;text-align:center;color:#8899bb;font-size:0.75rem;">
    Biz Maker — ビジネス共創プラットフォーム v7.0 — Growth Navigator &nbsp;|&nbsp; Phase 1 Enhanced &nbsp;|&nbsp; Powered by Streamlit
</div>
""", unsafe_allow_html=True)
